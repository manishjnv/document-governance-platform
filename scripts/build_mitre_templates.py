"""Generates the two downloadable MITRE templates (plan phase A11 piece 2):
styling only, every cell VALUE stays byte-identical to what ships today
(verified by the existing round-trip tests in test_mitre_ingest.py --
test_real_use_case_template_round_trips / test_real_environment_template_round_trips).

Prefer re-running this script over hand-editing the .xlsx files directly:
    cd apps/api && python ../../scripts/build_mitre_templates.py

Adds, per data sheet: a branded header-row fill (matches report_xlsx.py's
sheet() helper, Phase A11 piece 1) + white bold font, thin all-borders on
the header and example rows, and ~100 pre-formatted blank rows (border
only, no values) so customer-entered content lands in a visible grid.
The environment workbook's prose "Read Me" sheet is reproduced verbatim
(not a data-entry form) with its title row getting the same brand fill
for visual consistency with the rest of the product -- and its row
heights fixed while at it (an earlier ad hoc openpyxl `insert_rows()`
edit had left two rows' heights swapped and one row with no explicit
height at all; this generator sets every row's height explicitly).
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TEMPLATES_DIR = (
    Path(__file__).resolve().parent.parent / "apps" / "web" / "public" / "templates"
)

# Same brand color/font as report_xlsx.py's sheet() helper (Phase A11 piece 1)
# -- one consistent header treatment across the report AND the templates.
BRAND = "0057B8"
HEADER_FONT = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="8496AD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BLANK_ROWS = 100  # pre-formatted empty data rows per sheet


def _fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _data_sheet(wb, title, headers, example_rows, widths, *, first=False):
    """Header (brand fill + white bold + border) + example rows (border,
    values unchanged) + ~100 pre-formatted blank rows (border only) so the
    sheet reads as a fillable form/grid in Excel."""
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = _fill(BRAND)
        cell.border = BORDER
        cell.alignment = WRAP
    for row in example_rows:
        ws.append(list(row))
    n_cols = len(headers)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=n_cols):
        for cell in row:
            cell.border = BORDER
            cell.alignment = WRAP
    start = ws.max_row + 1
    for r in range(start, start + BLANK_ROWS):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).border = BORDER
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    return ws


def build_use_case_template() -> Path:
    wb = Workbook()
    _data_sheet(
        wb, "Rules",
        ["Use Case Name", "MITRE Technique IDs", "Detection Logic", "Description",
         "Log Source", "Status", "Severity", "Last Triggered"],
        [
            ["Suspicious PowerShell EncodedCommand", "T1059.001",
             "process=powershell.exe AND cmdline contains '-enc'",
             "Flags base64-encoded PowerShell, a common attacker execution pattern",
             "Sysmon", "Enabled", "High", "2026-07-28"],
            ["Legacy RDP Brute Force Watch", "T1110.001, T1021.001",
             "logon_type=10 AND failed_attempts > 20",
             "Disabled pending tuning -- kept for reference",
             "Windows Event Logs", "Disabled", "Medium", "never"],
            ["Unusual Outbound Data Transfer", None,
             "bytes_out > 500MB AND destination NOT IN known_list",
             "Broad exfiltration heuristic -- not yet mapped to ATT&CK",
             "Firewall", "Enabled", None, None],
            ["Suspicious Scheduled Task Creation", "T1053.005, T1059.001",
             "schtasks /create /tn * /tr *powershell*",
             "Scheduled task spawning PowerShell -- multi-technique example",
             "Sysmon", "Enabled", "Critical", "2026-08-01"],
        ],
        [38, 20, 48, 52, 14, 10, 14, 16],
        first=True,
    )
    path = TEMPLATES_DIR / "mitre-use-cases-template.xlsx"
    wb.save(path)
    return path


# (text, bold, height) -- verbatim from the shipped Read Me sheet.
_READ_ME_ROWS = [
    ("MITRE Environment Workbook", True, 30),
    (None, False, 8),
    ("This workbook is optional. Uploading it lets the assessment filter out "
     "techniques that don't apply to your environment and prioritize gaps "
     "by what you could realistically detect first. Skipping it still "
     "produces a coverage assessment -- just a conservative lower bound.",
     False, 30),
    (None, False, 8),
    ("Sheets in this workbook (all optional -- a missing sheet is noted as "
     "an assumption, never an error):", True, 30),
    (None, False, 8),
    ("Assets -- your platform/technology inventory (e.g. Windows, AWS, "
     "Azure AD). Used to mark techniques that don't apply to any platform "
     "you have (e.g. macOS-only techniques when you have no Mac fleet) as "
     "Not Applicable.", False, 30),
    ("Log Sources -- telemetry you already collect (e.g. Sysmon, "
     "CloudTrail). Used to prioritize gaps you could detect right now with "
     "data you already have. Optional health columns (Parser / Format, "
     "Normalized (Y/N), Last Event Seen) sharpen that prioritization -- a "
     "source that isn't normalized or hasn't seen recent events is treated "
     "as needing attention before you rely on it, not as fully ready.",
     False, 30),
    ('Tip: if one device sends more than one kind of log, list each log '
     'type as its own row in Log Sources -- e.g. "Infoblox - DNS logs" and '
     '"Infoblox - SSH logs" -- so each stream gets credited separately.',
     False, 30),
    ('Security Tooling -- products you own but may not have onboarded for '
     'logging yet. Used for the next-tier "onboard this, then detect" '
     "gaps.", False, 30),
    ("Crown Jewels -- free text describing your most critical assets (e.g. "
     '"customer database", "payment gateway"). Used only to nudge the '
     "order gaps are shown in when they're relevant to what you described "
     "-- it never changes a coverage percentage or a technique's status.",
     False, 30),
    (None, False, 8),
    ("Honesty note: this assessment never ingests raw logs and never verifies a "
     "specific field is present in your data. Everything here is metadata "
     "you tell us about your environment -- if something looks off, it's "
     "worth double-checking your own inventory, not just trusting this "
     "workbook.", False, 30),
]


def _build_read_me(wb) -> None:
    ws = wb.active
    ws.title = "Read Me"
    ws.column_dimensions["A"].width = 100
    for text, is_bold, height in _READ_ME_ROWS:
        ws.append([text])
        r = ws.max_row
        cell = ws.cell(row=r, column=1)
        cell.alignment = WRAP
        cell.font = TITLE_FONT if r == 1 else (BOLD if is_bold else Font())
        ws.row_dimensions[r].height = height
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = _fill(BRAND)


def build_environment_template() -> Path:
    wb = Workbook()
    _build_read_me(wb)

    _data_sheet(
        wb, "Assets", ["Platform"],
        [["Windows"], ["Linux"], ["Azure AD"], ["AWS"], ["Microsoft 365"]],
        [30],
    )
    _data_sheet(
        wb, "Log Sources",
        ["Source", "Present?", "Parser / Format", "Normalized (Y/N)", "Last Event Seen"],
        [
            ["Sysmon", "Yes", "Sysmon XML", "Yes", "2026-08-01"],
            ["Windows Event Logs", "Yes", None, None, None],
            ["CloudTrail", "Yes", "CloudTrail JSON", "No", "2025-03-15"],
        ],
        [30, 12, 18, 18, 18],
    )
    _data_sheet(
        wb, "Security Tooling", ["Tool"],
        [["CrowdStrike Falcon EDR"], ["Proofpoint Email Security"]],
        [30],
    )
    _data_sheet(
        wb, "Crown Jewels", ["Asset"],
        [["Customer database"], ["Payment processing service"]],
        [30],
    )

    path = TEMPLATES_DIR / "mitre-environment-template.xlsx"
    wb.save(path)
    return path


if __name__ == "__main__":
    uc_path = build_use_case_template()
    env_path = build_environment_template()
    print(f"wrote {uc_path}")
    print(f"wrote {env_path}")
