"""Generate the MITRE sample test-data kit under docs/sample/MITRE_Sample/.

Populated, upload-ready fixtures that exercise every MITRE-module path
(tagging ladder, coverage states, domain gating, roadmap buckets, threat
weighting, detection strength, trend compare, column wizard, exports).
Companion playbook: docs/sample/MITRE_Sample/README.md (hand-maintained).

Every technique ID is validated against the pinned ATT&CK dataset
(app.mitre.attack_data, v19.1) before writing, and every produced file is
round-tripped through the real ingest parsers afterwards. Re-run after any
ATT&CK re-pin:

    python scripts/generate_mitre_samples.py

(Needs the apps/api environment: openpyxl + the app package.)
"""

import csv
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "apps" / "api"))

from openpyxl import Workbook  # noqa: E402

from app.mitre.attack_data import DEFAULT  # noqa: E402
from app.mitre.ingest import (  # noqa: E402
    extract_tags,
    parse_environment_file,
    parse_use_case_file,
)
from app.mitre.keyword_tag import keyword_tag_rows  # noqa: E402

OUT = REPO / "docs" / "sample" / "MITRE_Sample"

# ---------------------------------------------------------------------------
# ID contract with the pinned dataset — the generator refuses to write a kit
# whose IDs don't behave exactly as the fixtures advertise.
# ---------------------------------------------------------------------------

EXPECTED_STATUS = {
    # customer-tagged valid rows
    "T1685.005": "ok",   # Clear Windows Event Logs (v19.1 successor of T1070.001)
    "T1003.001": "ok",   # LSASS Memory
    "T1547.001": "ok",   # Registry Run Keys / Startup Folder
    "T1059": "ok",       # Command and Scripting Interpreter (parent)
    "T1059.001": "ok",   # PowerShell (sub)
    "T0806": "ok",       # ICS: Brute Force I/O
    "T0809": "ok",       # ICS: Data Destruction
    "T1414": "ok",       # Mobile: Clipboard Data
    "T1407": "ok",       # Mobile: Download New Code at Runtime
    "T1558.003": "ok",   # Kerberoasting
    "T1566.001": "ok",   # Spearphishing Attachment
    "T1105": "ok",       # Ingress Tool Transfer
    "T1021.001": "ok",   # Remote Desktop Protocol
    # deliberately-bad rows
    "T1562.001": "remapped",    # revoked in v19.1 -> T1685
    "T9999": "unknown",         # well-formed, not in dataset
    "T1026": "deprecated",      # Multiband Communication
    # v2-only additions
    "T1078": "ok",       # Valid Accounts
    "T1071.001": "ok",   # Web Protocols
    "T1486": "ok",       # Data Encrypted for Impact
}

# Untagged rows the deterministic keyword pre-pass must map (row index ->
# technique the emitted set must contain), and rows it must NOT touch.
KEYWORD_EXPECT = {
    18: "T1003.001",   # mimikatz / sekurlsa
    19: "T1053.005",   # schtasks
    20: "T1059.001",   # powershell / -enc
    21: "T1218.011",   # rundll32
    22: "T1047",       # wmic
    23: "T1490",       # vssadmin delete / delete shadows
    24: "T1685.005",   # wevtutil cl (alias targets T1070.001 -> remapped)
    25: "T1105",       # certutil
    26: "T1482",       # nltest
    27: "T1567.002",   # rclone
}
AI_RESIDUE_ROWS = {28, 29, 30}  # must survive the keyword pass unmatched

USE_CASE_HEADERS = [
    "Use Case Name", "MITRE Technique IDs", "Detection Logic",
    "Description", "Log Source", "Status", "Feature Exercised",
]

# (name, tags, logic, description, log_source, status, feature-exercised note)
PRIMARY_ROWS = [
    ("Windows Event Log Cleared",
     "T1685.005",
     "index=wineventlog EventCode=1102 | stats count by host, user",
     "Alerts when the Security event log is cleared on a Windows host.",
     "Windows Security Events", "Enabled",
     "customer-tagged valid ID -> covered"),
    ("LSASS Memory Access via Comsvcs",
     "T1003.001",
     'index=edr process_name=rundll32.exe cmdline="*comsvcs.dll MiniDump*"',
     "Detects LSASS memory dumping through the comsvcs.dll MiniDump export.",
     "EDR Telemetry", "Enabled",
     "description AND logic populated -> Phase 7 logic persistence + Phase 12 telemetry match"),
    ("Registry Run Key Persistence - With Logic",
     "T1547.001",
     'index=sysmon EventCode=13 TargetObject="*\\\\CurrentVersion\\\\Run*"',
     "Registry run-key writes captured by Sysmon.",
     "Sysmon", "Enabled",
     "logic present -> stronger detection strength (Phase 12 pair A)"),
    ("Registry Run Key Persistence - Baseline",
     "T1547.001",
     "",
     "Placeholder rule imported from a legacy SIEM without query text.",
     "Sysmon", "Enabled",
     "same technique, logic absent -> weaker detection strength (Phase 12 pair B)"),
    ("Suspicious Script Interpreter Spawn",
     "T1059",
     "index=edr event=process_start interpreter IN (sh, bash, python)",
     "Broad interpreter-spawn detection at the parent-technique level.",
     "EDR Telemetry", "Enabled",
     "parent technique row -> sub-technique rollup (with row 7)"),
    ("PowerShell Suspicious Flags",
     "T1059.001",
     'index=edr process=powershell.exe cmdline IN ("*-nop*", "*bypass*")',
     "PowerShell launched with suspicious flags.",
     "EDR Telemetry", "Enabled",
     "sub-technique of row 6 -> rollup"),
    ("PLC Brute Force I/O Writes",
     "T0806",
     "ot_historian: forced_write_count > threshold by plc_id",
     "Repeated forced I/O writes against PLC points.",
     "OT Telemetry", "Enabled",
     "ICS technique -> ICS domain coverage when environment has OT assets"),
    ("Historian Data Destruction",
     "T0809",
     "ot_historian: bulk delete of process tags",
     "Bulk deletion of historian process data.",
     "OT Telemetry", "Enabled",
     "second ICS technique"),
    ("Mobile Clipboard Scraping App",
     "T1414",
     "mdm: app with clipboard-read entitlement anomaly",
     "MDM flags apps abusing clipboard read on managed devices.",
     "Intune MDM", "Enabled",
     "Mobile technique -> Mobile domain coverage when fleet is managed"),
    ("Mobile App Downloads Code at Runtime",
     "T1407",
     "mdm: app fetches executable payload post-install",
     "Managed-device app pulling new code after install.",
     "Intune MDM", "Enabled",
     "second Mobile technique"),
    ("Kerberos RC4 Service Ticket Anomaly",
     "T1558.003",
     "index=wineventlog EventCode=4769 encryption_type=0x17 | rare service_name",
     "RC4 TGS requests indicating possible Kerberoasting.",
     "Windows Security Events", "Enabled",
     "enabled here; DISABLED in v2 -> shows REGRESSED in trend compare"),
    ("Spearphishing Attachment Detonated",
     "T1566.001",
     "email_gw: attachment sandbox verdict = malicious",
     "Sandbox-detonated malicious attachment from inbound mail.",
     "Email Security Gateway", "Enabled",
     "customer-tagged valid ID -> covered"),
    ("Malicious Script Downloads Payload",
     "T1059.001, T1105",
     'index=edr process=powershell.exe dest_url=* file_download=true',
     "Script interpreter fetching a remote payload.",
     "EDR Telemetry", "Enabled",
     "MULTIPLE IDs in one cell -> multi-tag extraction"),
    ("RDP Lateral Movement Watch",
     "T1021.001",
     "index=wineventlog EventCode=4624 LogonType=10 | unusual source",
     "Interactive RDP logons from unusual sources.",
     "Windows Security Events", "Disabled",
     "customer-tagged but DISABLED -> partial (default policy); Enabled in v2 -> newly covered"),
    ("Defender Tampering Alert",
     "T1562.001",
     "index=edr defender_status=disabled by non-admin",
     "Security tooling disabled outside change control.",
     "EDR Telemetry", "Enabled",
     "REVOKED ID in v19.1 -> auto-remap to T1685 + assumption note"),
    ("Legacy Rule With Broken Tags",
     "T9999, banana",
     "index=legacy correlation id 42",
     "Migrated rule whose tag field was corrupted.",
     "Legacy SIEM", "Enabled",
     "invalid tags: T9999 -> unknown; 'banana' dropped by the tag regex at ingest"),
    ("Multiband Communication Detection",
     "T1026",
     "ndr: multi-channel C2 correlation",
     "Legacy multi-band C2 rule tagged with a retired technique.",
     "NDR Sensor", "Enabled",
     "DEPRECATED ID in v19.1 -> deprecated note"),
    # -- untagged, keyword-taggable (rows 19-28 on-sheet) ------------------
    ("Mimikatz Execution",
     "",
     'index=edr cmdline="*mimikatz*" OR cmdline="*sekurlsa::logonpasswords*"',
     "Known credential-theft tool execution.",
     "EDR Telemetry", "Enabled",
     "keyword pre-pass: mimikatz -> T1003.001 (no LLM)"),
    ("Scheduled Task Created via CLI",
     "",
     'index=wineventlog process=schtasks.exe cmdline="*schtasks /create*"',
     "Task registration from the command line.",
     "Windows Security Events", "Enabled",
     "keyword: schtasks -> T1053.005"),
    ("Encoded PowerShell Command",
     "",
     'index=edr cmdline="powershell.exe -enc *"',
     "Base64-encoded command execution.",
     "EDR Telemetry", "Enabled",
     "keyword: powershell / -enc -> T1059.001"),
    ("Rundll32 Proxy Launch",
     "",
     "index=sysmon process=rundll32.exe parent!=msiexec.exe",
     "Signed-binary proxy execution via rundll32.",
     "Sysmon", "Enabled",
     "keyword: rundll32 -> T1218.011"),
    ("WMIC Remote Process Launch",
     "",
     'index=edr cmdline="wmic /node:* process call create *"',
     "Remote process creation over WMI.",
     "EDR Telemetry", "Enabled",
     "keyword: wmic -> T1047"),
    ("Volume Shadow Copy Removal",
     "",
     'index=edr cmdline="vssadmin delete shadows /all /quiet"',
     "Backup destruction ahead of ransomware.",
     "EDR Telemetry", "Enabled",
     "keyword: vssadmin delete -> T1490"),
    ("Event Log Clear via Wevtutil",
     "",
     'index=edr cmdline="wevtutil cl Security"',
     "Log clearing from the command line.",
     "EDR Telemetry", "Enabled",
     "keyword: wevtutil cl -> T1070.001, remapped to T1685.005"),
    ("Certutil Remote Download",
     "",
     'index=edr cmdline="certutil -urlcache -split -f http*"',
     "LOLBin file download.",
     "EDR Telemetry", "Enabled",
     "keyword: certutil -> T1105"),
    ("Nltest Domain Trust Enumeration",
     "",
     'index=edr cmdline="nltest /domain_trusts"',
     "Domain trust discovery from an endpoint.",
     "EDR Telemetry", "Enabled",
     "keyword: nltest -> T1482"),
    ("Rclone Exfiltration",
     "",
     'index=edr process=rclone.exe cmdline="*copy* *remote:*"',
     "Bulk copy to attacker-controlled cloud storage.",
     "EDR Telemetry", "Enabled",
     "keyword: rclone -> T1567.002"),
    # -- untagged, genuinely non-keywordable (AI residue) ------------------
    ("Impossible Travel Sign-In",
     "",
     "SigninLogs | geo distance between consecutive sign-ins vs elapsed time",
     "Sign-ins from two distant locations within an hour.",
     "Entra ID Sign-In Logs", "Enabled",
     "untagged, non-keywordable -> AI tagging residue (or 'unmapped' if no LLM key); tagged T1078 in v2"),
    ("Periodic Outbound Callback Pattern",
     "",
     "ndr: fixed-interval outbound connections with low jitter per host",
     "Regular low-jitter callbacks flagged by the UEBA model.",
     "NDR Sensor", "Enabled",
     "untagged, non-keywordable -> AI residue; tagged T1071.001 in v2"),
    ("DLP Bulk Upload Spike",
     "",
     "dlp: outbound upload volume > 5x user baseline in 24h",
     "Sudden large data uploads relative to the user's baseline.",
     "DLP", "Enabled",
     "untagged, non-keywordable -> AI residue (or 'unmapped')"),
]

V2_NEW_ROW = (
    "Ransomware Mass Encryption Behavior",
    "T1486",
    "index=edr file_rename_rate > threshold AND entropy_delta high",
    "Mass file encryption behavior on endpoints.",
    "EDR Telemetry", "Enabled",
    "NEW in v2 -> newly covered in trend compare",
)

MESSY_HEADERS = [
    "Rule Title", "Ref Codes", "Payload", "Meaning", "Feed", "Mode",
    "Feature Exercised",
]
# Column-wizard mapping to enter for usecases_messy_headers.xlsx:
MESSY_OVERRIDE = {
    "name": 0, "tags": 1, "logic": 2, "description": 3,
    "log_source": 4, "enabled": 5,
}


def v2_rows():
    rows = []
    for i, row in enumerate(PRIMARY_ROWS, start=1):
        row = list(row)
        if i == 11:   # Kerberoasting -> disabled: REGRESSED
            row[5] = "Disabled"
            row[6] = "was Enabled in primary -> REGRESSED in trend compare"
        elif i == 14:  # RDP watch -> enabled: newly covered
            row[5] = "Enabled"
            row[6] = "was Disabled in primary -> NEWLY COVERED in trend compare"
        elif i == 28:  # impossible travel now customer-tagged
            row[1] = "T1078"
            row[6] = "untagged in primary, tagged T1078 here -> NEWLY COVERED"
        elif i == 29:  # callback pattern now customer-tagged
            row[1] = "T1071.001"
            row[6] = "untagged in primary, tagged T1071.001 here -> NEWLY COVERED"
        rows.append(tuple(row))
    rows.append(V2_NEW_ROW)
    return rows


# ---------------------------------------------------------------------------
# Environment workbooks. Sheet names must match ingest.SHEET_SYNONYMS; the
# first row of each sheet is a header ingest skips; an optional second
# column is the present?-flag (a recognizable "No" skips the row).
# ---------------------------------------------------------------------------

ENV_FULL = {
    "Assets": [
        ("Asset", "Present"),
        ("Windows Server 2022 fleet", "Yes"),
        ("Windows 11 endpoints", "Yes"),
        ("Ubuntu Linux servers", "Yes"),
        ("macOS laptops", "Yes"),
        ("Microsoft Entra ID tenant", "Yes"),
        ("AWS EC2 workloads", "Yes"),
        ("Microsoft 365 (Exchange Online, SharePoint)", "Yes"),
        ("Kubernetes clusters (EKS)", "Yes"),
        ("Palo Alto firewalls", "Yes"),
        ("OT/SCADA PLC segment", "Yes"),          # gates ICS applicable
        ("Managed mobile fleet via Intune MDM", "Yes"),  # gates Mobile
        ("Android corporate devices", "Yes"),
        ("iPhone / iPad fleet", "Yes"),
        ("VMware ESXi estate", "No"),   # present=No -> skipped (proves the flag)
        ("Mainframe z/OS", "Yes"),      # unmatched -> assumption note
    ],
    "Log Sources": [
        ("Log Source",),
        ("Windows Event Logs",), ("Sysmon",), ("CrowdStrike EDR telemetry",),
        ("Entra ID sign-in logs",), ("DNS query logs",), ("AWS CloudTrail",),
        ("M365 Unified Audit Log",), ("Intune MDM logs",),
        ("OT historian telemetry",), ("Palo Alto firewall logs",),
        ("NDR sensor metadata",),
    ],
    "Security Tooling": [
        ("Tool",),
        ("CrowdStrike Falcon EDR",), ("Proofpoint Email Security",),
        ("Darktrace NDR",), ("Microsoft Purview DLP",), ("CyberArk PAM",),
        ("Okta MFA",),
    ],
    "Crown Jewels": [
        ("Crown Jewel",),
        ("Payment processing platform",), ("Customer PII database",),
        ("Active Directory / Entra ID",), ("SCADA control network",),
        ("Source code repositories",),
    ],
}

ENV_WINDOWS_ONLY = {
    "Assets": [
        ("Asset",),
        ("Windows Server 2019 fleet",),
        ("Windows 10/11 endpoints",),
        ("Active Directory domain controllers",),
        ("Microsoft Entra ID tenant",),
    ],
    "Log Sources": [
        ("Log Source",),
        ("Windows Event Logs",), ("Sysmon",), ("AD authentication logs",),
    ],
    "Security Tooling": [
        ("Tool",),
        ("Microsoft Defender for Endpoint",),
    ],
    "Crown Jewels": [
        ("Crown Jewel",),
        ("Active Directory",), ("Finance file server",),
    ],
}


# ---------------------------------------------------------------------------
# Writers
# ---------------------------------------------------------------------------

def write_use_case_xlsx(path, rows, headers=USE_CASE_HEADERS):
    wb = Workbook()
    ws = wb.active
    ws.title = "Use Cases"
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    for col, width in zip("ABCDEFG", (42, 22, 60, 50, 24, 10, 60)):
        ws.column_dimensions[col].width = width
    wb.save(path)


def write_env_xlsx(path, sheets):
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(list(row))
        ws.column_dimensions["A"].width = 48
    wb.save(path)


def write_csv(path, rows):
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(USE_CASE_HEADERS)
        writer.writerows(rows)


def write_minimal_pdf(path, lines):
    """One-page text PDF, stdlib only (exercises the Phase 2 extraction
    path; the structured parsers reject pdf by design).
    ponytail: hand-rolled minimal PDF instead of a reportlab dependency."""

    def esc(s):
        return s.replace("\\", r"\\").replace("(", r"\(").replace(")", r"\)")

    stream = "BT /F1 10 Tf 50 790 Td 13 TL\n"
    stream += "".join(f"({esc(line)}) Tj T*\n" for line in lines)
    stream += "ET"
    stream_b = stream.encode("latin-1", "replace")

    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 842] "
        b"/Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>",
        b"<< /Length %d >>\nstream\n%s\nendstream" % (len(stream_b), stream_b),
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
    ]
    out = bytearray(b"%PDF-1.4\n")
    offsets = []
    for i, body in enumerate(objects, start=1):
        offsets.append(len(out))
        out += b"%d 0 obj\n%s\nendobj\n" % (i, body)
    xref_at = len(out)
    out += b"xref\n0 %d\n0000000000 65535 f \n" % (len(objects) + 1)
    for off in offsets:
        out += b"%010d 00000 n \n" % off
    out += (
        b"trailer\n<< /Size %d /Root 1 0 R >>\nstartxref\n%d\n%%%%EOF\n"
        % (len(objects) + 1, xref_at)
    )
    path.write_bytes(bytes(out))


PDF_LINES = [
    "ACME Corp - SOC Detection Rules (scanned export, page 1 of 1)",
    "",
    "Rule: Suspicious PowerShell Download Cradle",
    "Technique: T1059.001. Logic: powershell.exe with Net.WebClient in the",
    "command line, observed by EDR telemetry. Status: enabled.",
    "",
    "Rule: RDP Brute Force Burst",
    "Technique: T1110. Logic: more than 20 failed LogonType=10 attempts from",
    "one source in 5 minutes. Status: enabled.",
    "",
    "Rule: Suspicious OAuth App Consent",
    "No technique tag recorded. Logic: new service principal granted",
    "mail.read across the tenant outside change control. Status: enabled.",
]


# ---------------------------------------------------------------------------
# Validation + round-trip
# ---------------------------------------------------------------------------

def validate_ids():
    print(f"pinned ATT&CK version: {DEFAULT.version}")
    assert DEFAULT.version == "19.1", "kit was designed against v19.1 - re-check EXPECTED_STATUS"
    all_rows = PRIMARY_ROWS + [V2_NEW_ROW] + [("x", "T1078", "", "", "", "", ""),
                                              ("x", "T1071.001", "", "", "", "", "")]
    used = {tag for row in all_rows for tag in extract_tags(row[1])}
    unexpected = used - set(EXPECTED_STATUS)
    assert not unexpected, f"IDs used but not in EXPECTED_STATUS: {unexpected}"
    for tid, want in EXPECTED_STATUS.items():
        canonical, status = DEFAULT.resolve(tid)
        assert status == want, f"{tid}: expected {want}, dataset says {status} ({canonical})"
        print(f"  {tid:<10} -> {status:<10} {canonical or '-'}")
    # domain spot checks for the ICS/Mobile rows
    for tid, domain in (("T0806", "ics"), ("T0809", "ics"),
                        ("T1414", "mobile"), ("T1407", "mobile")):
        assert DEFAULT.get(tid)["domain"] == domain, f"{tid} not in {domain}"
    print("ID contract holds.\n")


def tag_split(parsed):
    counts = {"customer_tagged": 0, "untagged": 0}
    statuses = {}
    for row in parsed["rows"]:
        if not row["tags"]:
            counts["untagged"] += 1
            continue
        counts["customer_tagged"] += 1
        for tag in row["tags"]:
            _, status = DEFAULT.resolve(tag)
            statuses[status] = statuses.get(status, 0) + 1
    return counts, statuses


def roundtrip_use_cases(path, file_type, expect_rows, column_override=None,
                        check_keyword=False):
    parsed = parse_use_case_file(path.read_bytes(), file_type,
                                 column_override=column_override)
    counts, statuses = tag_split(parsed)
    print(f"[{path.name}]{' (with wizard override)' if column_override else ''}")
    print(f"  sheet={parsed['sheet']}  columns={parsed['columns']}")
    print(f"  rows={parsed['row_count']}  split={counts}  tag statuses={statuses}")
    for w in parsed["warnings"]:
        print(f"  warning: {w}")
    assert parsed["row_count"] == expect_rows, (
        f"{path.name}: {parsed['row_count']} rows, expected {expect_rows}")

    if check_keyword:
        untagged = [r for r in parsed["rows"] if not r["tags"]]
        matched = keyword_tag_rows(untagged)
        ref_of = {i + 1: r["row_ref"] for i, r in enumerate(parsed["rows"])}
        for idx, want in KEYWORD_EXPECT.items():
            got = [m["technique_id"] for m in matched.get(ref_of[idx], [])]
            assert want in got, f"row {idx}: keyword pass gave {got}, expected {want}"
        for idx in AI_RESIDUE_ROWS:
            assert ref_of[idx] not in matched, (
                f"row {idx} unexpectedly keyword-matched: {matched.get(ref_of[idx])}")
        residue = [r["row_ref"] for r in untagged if r["row_ref"] not in matched]
        print(f"  keyword-tagged {len(matched)}/{len(untagged)} untagged rows; "
              f"AI residue: {residue}")
    return parsed


def roundtrip_environment(path, expect_platforms, expect_ics, expect_mobile):
    parsed = parse_environment_file(path.read_bytes(), "xlsx")
    env = parsed["environment"]
    print(f"[{path.name}]")
    print(f"  sheets={parsed['sheets_found']}")
    print(f"  platforms={env['platforms']}")
    print(f"  has_ics={env['has_ics_assets']}  has_mobile={env['has_managed_mobile']}")
    print(f"  log_sources={len(parsed['log_sources'])}  tooling={len(parsed['tooling'])}  "
          f"crown_jewels={len(parsed['crown_jewels'])}")
    for a in parsed["assumptions"]:
        print(f"  assumption: {a}")
    assert set(env["platforms"]) == expect_platforms, (
        f"{path.name}: platforms {env['platforms']}, expected {sorted(expect_platforms)}")
    assert env["has_ics_assets"] is expect_ics
    assert env["has_managed_mobile"] is expect_mobile
    assert len(parsed["sheets_found"]) == 4, "not all four sheets recognized"


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    validate_ids()

    v2 = v2_rows()
    write_use_case_xlsx(OUT / "usecases_primary.xlsx", PRIMARY_ROWS)
    write_use_case_xlsx(OUT / "usecases_v2_improved.xlsx", v2)
    write_use_case_xlsx(OUT / "usecases_messy_headers.xlsx", PRIMARY_ROWS,
                        headers=MESSY_HEADERS)
    write_env_xlsx(OUT / "environment_full.xlsx", ENV_FULL)
    write_env_xlsx(OUT / "environment_windows_only.xlsx", ENV_WINDOWS_ONLY)
    write_csv(OUT / "usecases_primary.csv", PRIMARY_ROWS)
    write_minimal_pdf(OUT / "usecases_scanned.pdf", PDF_LINES)
    # .xls deliberately skipped: no xls writer installed (xlrd reads only).

    print("--- round-trip through the real ingest parsers ---")
    roundtrip_use_cases(OUT / "usecases_primary.xlsx", "xlsx",
                        expect_rows=len(PRIMARY_ROWS), check_keyword=True)
    roundtrip_use_cases(OUT / "usecases_v2_improved.xlsx", "xlsx",
                        expect_rows=len(v2))
    roundtrip_use_cases(OUT / "usecases_primary.csv", "csv",
                        expect_rows=len(PRIMARY_ROWS))

    # messy headers: auto-detection must find the name column ("Rule Title"
    # is a known synonym) but MISS the tags column -> wizard territory...
    auto = roundtrip_use_cases(OUT / "usecases_messy_headers.xlsx", "xlsx",
                               expect_rows=len(PRIMARY_ROWS))
    assert "tags" not in auto["columns"], (
        "messy file's tags column was auto-detected - rename 'Ref Codes'")
    # ...and the wizard override must recover the full mapping.
    fixed = roundtrip_use_cases(OUT / "usecases_messy_headers.xlsx", "xlsx",
                                expect_rows=len(PRIMARY_ROWS),
                                column_override=MESSY_OVERRIDE)
    assert set(fixed["columns"]) == set(MESSY_OVERRIDE)

    roundtrip_environment(
        OUT / "environment_full.xlsx",
        expect_platforms={"Windows", "Linux", "macOS", "Identity Provider",
                          "IaaS", "Office Suite", "Containers",
                          "Network Devices", "Android", "iOS"},
        expect_ics=True, expect_mobile=True)
    roundtrip_environment(
        OUT / "environment_windows_only.xlsx",
        expect_platforms={"Windows", "Identity Provider"},
        expect_ics=False, expect_mobile=False)

    print("\nAll files written to", OUT)
    print("All round-trip assertions passed.")


if __name__ == "__main__":
    main()
