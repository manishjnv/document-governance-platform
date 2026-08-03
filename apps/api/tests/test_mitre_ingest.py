"""Ingest unit tests for the MITRE assessment module (Phase 1).

Fixture files are built in-memory with openpyxl / plain CSV — no checked-in
binaries. (.xls reading via xlrd is untested here: no xls writer is
installed; the code path is symmetric with xlsx.)
"""

import io
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.mitre import ingest
from app.mitre.ingest import IngestError, parse_use_case_file, parse_environment_file

TEMPLATES_DIR = (
    Path(__file__).resolve().parents[3] / "apps" / "web" / "public" / "templates"
)


def _xlsx(rows, sheet_name="Sheet1", extra_sheets=()) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    for name, extra_rows in extra_sheets:
        extra = wb.create_sheet(name)
        for row in extra_rows:
            extra.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


TEMPLATE_HEADERS = [
    "Use Case Name", "MITRE Technique(s)", "Detection Logic",
    "Description", "Log Source", "Status",
]


def test_template_layout_happy_path():
    """Plan phase A6 old-layout regression: a dump with NO Severity/Last
    Triggered columns must still detect exactly the pre-A6 column set and
    produce rows with those two new fields present but None."""
    content = _xlsx([
        TEMPLATE_HEADERS,
        ["Suspicious PowerShell", "T1059.001", "process where ...", "PS abuse", "Sysmon", "Enabled"],
        ["RDP Brute Force", "T1110, T1021.001", "auth failures > 20", "", "WinEventLog", "Disabled"],
        ["Untagged rule", "", "some query", "", "", "Enabled"],
    ])
    parsed = parse_use_case_file(content, "xlsx")
    assert parsed["row_count"] == 3
    assert set(parsed["columns"]) == {"name", "tags", "logic", "description", "log_source", "enabled"}
    r1, r2, r3 = parsed["rows"]
    assert r1 == {
        "row_ref": "Sheet1:2", "name": "Suspicious PowerShell",
        "description": "PS abuse", "log_source": "Sysmon", "enabled": True,
        "tags": ["T1059.001"], "logic": "process where ...",
        "severity": None, "last_triggered": None,
    }
    assert r2["tags"] == ["T1110", "T1021.001"]
    assert r2["enabled"] is False
    assert r3["tags"] == [] and r3["enabled"] is True


def test_messy_headers_detected_below_title_rows():
    content = _xlsx([
        ["SOC Detection Inventory — Export 2026"],
        [],
        ["Rule", "TTPs", "Query", "State"],
        ["Kerberoasting watch", "t1558.003", "spn requests...", "active"],
    ])
    parsed = parse_use_case_file(content, "xlsx")
    assert parsed["row_count"] == 1
    row = parsed["rows"][0]
    assert row["row_ref"] == "Sheet1:4"
    assert row["tags"] == ["T1558.003"]  # lowercase tag normalized
    assert row["enabled"] is True


def test_csv_happy_path():
    content = "Rule Name,MITRE ID,Status\r\nDNS tunnel detect,T1071.004,enabled\r\n".encode()
    parsed = parse_use_case_file(content, "csv")
    assert parsed["row_count"] == 1
    assert parsed["rows"][0]["row_ref"] == "csv:2"
    assert parsed["rows"][0]["tags"] == ["T1071.004"]


def test_empty_file_rejected():
    with pytest.raises(IngestError, match="empty"):
        parse_use_case_file(b"", "xlsx")


def test_no_name_column_rejected_pointing_at_template():
    content = _xlsx([["foo", "bar"], ["a", "b"]])
    with pytest.raises(IngestError, match="template"):
        parse_use_case_file(content, "xlsx")


def test_row_cap_enforced():
    lines = ["Rule Name,MITRE ID"] + [f"rule {i},T1059" for i in range(ingest.MAX_USE_CASE_ROWS + 1)]
    with pytest.raises(IngestError, match="row limit|5,000"):
        parse_use_case_file("\n".join(lines).encode(), "csv")


def test_pdf_docx_rejected_with_phase2_message():
    for file_type in ("pdf", "docx"):
        with pytest.raises(IngestError, match="XLSX template"):
            parse_use_case_file(b"%PDF-1.4 whatever", file_type)


def test_environment_workbook_parsing():
    content = _xlsx(
        [["Platform"], ["Windows Server 2019"], ["Ubuntu Linux"], ["Azure AD"],
         ["AWS"], ["OT/SCADA network"], ["Cisco IOS switches"], ["SAP ERP"]],
        sheet_name="Assets",
        extra_sheets=[("Log Sources", [["Source"], ["Sysmon"], ["CloudTrail"]])],
    )
    parsed = parse_environment_file(content, "xlsx")
    env = parsed["environment"]
    assert env["inventory_provided"] is True
    assert set(env["platforms"]) == {"Windows", "Linux", "Identity Provider", "IaaS", "Network Devices"}
    assert env["has_ics_assets"] is True          # OT/SCADA row
    assert env["has_managed_mobile"] is False
    assert parsed["log_sources"] == ["Sysmon", "CloudTrail"]
    assert parsed["log_source_health"] == {}  # Phase A6 old-layout: no health columns -> no-op
    assert parsed["sheets_found"] == {"assets": "Assets", "log_sources": "Log Sources"}
    # missing sheets + unmapped assets become assumption lines
    assert any("Security Tooling" in a for a in parsed["assumptions"])
    assert any("Crown Jewels" in a for a in parsed["assumptions"])
    assert any("SAP ERP" in a for a in parsed["assumptions"])
    # Phase 14g: per-entry evidence trail (additive output)
    by_entry = {i["entry"]: i for i in parsed["interpretations"]}
    assert by_entry["Windows Server 2019"]["interpretation"] == "counted as platform Windows"
    assert "enabled the ICS/OT matrix" in by_entry["OT/SCADA network"]["interpretation"]
    assert by_entry["SAP ERP"]["interpretation"] == (
        "not recognized — ignored for platform filtering"
    )
    assert by_entry["Sysmon"]["sheet"] == "Log Sources"
    assert "short-term" in by_entry["Sysmon"]["interpretation"]


def test_environment_present_no_rows_recorded_as_skipped():
    """Phase 14g: a Present=No row is excluded from parsing (unchanged) AND
    shows up in the evidence trail as skipped."""
    content = _xlsx(
        [["Asset", "Present?"], ["Windows fleet", "Yes"], ["VMware ESXi estate", "No"]],
        sheet_name="Assets",
    )
    parsed = parse_environment_file(content, "xlsx")
    assert parsed["environment"]["platforms"] == ["Windows"]
    by_entry = {i["entry"]: i for i in parsed["interpretations"]}
    assert by_entry["VMware ESXi estate"]["interpretation"] == (
        "skipped — you marked it Present = No"
    )


def test_environment_mobile_detection_and_missing_assets_sheet():
    content = _xlsx(
        [["Source"], ["Sysmon"]],
        sheet_name="Log Sources",
        extra_sheets=[("Tooling", [["Tool"], ["Intune MDM"], ["CrowdStrike EDR"]])],
    )
    parsed = parse_environment_file(content, "xlsx")
    assert parsed["environment"]["inventory_provided"] is False
    assert any("Assets" in a for a in parsed["assumptions"])

    content = _xlsx([["Asset"], ["Android"], ["iPhone fleet"]], sheet_name="Assets")
    parsed = parse_environment_file(content, "xlsx")
    env = parsed["environment"]
    assert set(env["platforms"]) == {"Android", "iOS"}
    assert env["has_managed_mobile"] is True


def test_environment_must_be_excel():
    with pytest.raises(IngestError, match="Excel"):
        parse_environment_file(b"a,b,c", "csv")


# ---------------------------------------------------------------------------
# workbook-wide budgets (2026-08-01 adversarial review, blocking finding #2)
# ---------------------------------------------------------------------------


def test_workbook_sheet_cap():
    wb = Workbook()
    for i in range(ingest.MAX_SHEETS + 1):
        wb.create_sheet(f"extra{i}")
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(IngestError, match="sheets"):
        ingest._xlsx_grids(buf.getvalue())


def test_workbook_cumulative_row_cap():
    wb = Workbook()
    half = ingest.MAX_TOTAL_ROWS // 2 + 100
    for title in ("a", "b"):
        ws = wb.create_sheet(title)
        for _ in range(half):
            ws.append(["x"])
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(IngestError, match="overall limit"):
        ingest._xlsx_grids(buf.getvalue())


def test_workbook_row_width_cap():
    wb = Workbook()
    wb.active.append(["c"] * (ingest.MAX_ROW_CELLS + 1))
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(IngestError, match="wider than"):
        ingest._xlsx_grids(buf.getvalue())


# --- Phase 9: column override + preview headers/samples ---

def test_column_override_replaces_detection():
    content = _xlsx([
        ["Rule Name", "Ref Codes", "Query"],
        ["Encoded PowerShell", "T1059.001", "proc = powershell"],
    ])
    auto = parse_use_case_file(content, "xlsx")
    assert "tags" not in auto["columns"]  # 'Ref Codes' isn't a synonym
    assert auto["headers"] == ["Rule Name", "Ref Codes", "Query"]
    assert auto["sample_rows"] == [["Encoded PowerShell", "T1059.001", "proc = powershell"]]

    overridden = parse_use_case_file(
        content, "xlsx", column_override={"name": 0, "tags": 1, "logic": 2}
    )
    assert overridden["columns"] == {"name": 0, "tags": 1, "logic": 2}
    assert overridden["rows"][0]["tags"] == ["T1059.001"]
    assert overridden["rows"][0]["logic"] == "proc = powershell"


def test_column_override_validation_errors():
    content = _xlsx([["Rule Name", "Ref Codes"], ["r1", "T1110"]])
    for bad, needle in [
        ({"name": 5}, "out of range"),
        ({"name": True}, "out of range"),          # bools are not indexes
        ({"nope": 0, "name": 1}, "Unknown field"),
        ({"tags": 1}, "name column is required"),
        ({}, "non-empty object"),
        ({"name": 0, "tags": 0}, "one field"),
    ]:
        with pytest.raises(IngestError, match=needle):
            parse_use_case_file(content, "xlsx", column_override=bad)


def test_csv_grid_enforces_row_and_width_caps():
    wide = ("name," + ",".join(["x"] * ingest.MAX_ROW_CELLS)).encode()
    with pytest.raises(IngestError, match="wider than"):
        ingest._csv_grid(wide)
    tall = ("name\n" + "r\n" * (ingest.MAX_TOTAL_ROWS + 1)).encode()
    with pytest.raises(IngestError, match="overall limit"):
        ingest._csv_grid(tall)


# --- Phase 6: widened header/sheet/platform synonym sets ---

def test_splunk_style_headers_detected():
    content = _xlsx([
        ["Correlation Search Name", "ATT&CK ID", "SPL Query", "Objective", "Deployment Status", "Source Type"],
        ["Encoded PowerShell", "T1059.001", "| tstats ...", "catch encodedcommand", "Enabled", "wineventlog"],
    ])
    parsed = parse_use_case_file(content, "xlsx")
    assert set(parsed["columns"]) == {"name", "tags", "logic", "description", "enabled", "log_source"}
    row = parsed["rows"][0]
    assert row["name"] == "Encoded PowerShell"
    assert row["tags"] == ["T1059.001"]
    assert row["enabled"] is True
    assert row["log_source"] == "wineventlog"


def test_sentinel_style_headers_detected():
    content = _xlsx([
        ["Analytic Name", "MITRE_TTP", "KQL Query", "Is_Enabled", "Log_Type"],
        ["OAuth consent grant", "T1528", "AuditLogs | where ...", "true", "AuditLogs"],
    ])
    parsed = parse_use_case_file(content, "xlsx")
    assert set(parsed["columns"]) == {"name", "tags", "logic", "enabled", "log_source"}
    assert parsed["rows"][0]["tags"] == ["T1528"]
    assert parsed["rows"][0]["enabled"] is True


def test_widened_environment_sheet_and_platform_synonyms():
    content = _xlsx(
        [["Asset"], ["Palo Alto firewalls"], ["SUSE Linux estate"], ["AKS clusters"],
         ["Duo MFA"], ["iPadOS tablets"], ["Modbus PLC network"]],
        sheet_name="Asset List",
        extra_sheets=[
            ("SIEM Sources", [["Source"], ["Sysmon"]]),
            ("Security Products", [["Tool"], ["CrowdStrike"]]),
            ("High Value Assets", [["Item"], ["Payment gateway"]]),
        ],
    )
    parsed = parse_environment_file(content, "xlsx")
    assert set(parsed["sheets_found"]) == {"assets", "log_sources", "tooling", "crown_jewels"}
    env = parsed["environment"]
    assert set(env["platforms"]) == {
        "Network Devices", "Linux", "Containers", "Identity Provider", "iOS",
    }
    assert env["has_managed_mobile"] is True   # iPadOS fleet
    assert env["has_ics_assets"] is True       # Modbus marker
    assert parsed["tooling"] == ["CrowdStrike"]
    assert parsed["crown_jewels"] == ["Payment gateway"]


# --- Phase A10 piece 1: device-level platform synonyms (the "Infoblox problem") ---

def test_a10_photon_infoblox_rubrik_platform_synonyms():
    content = _xlsx(
        [["Asset"], ["Photon OS appliances"], ["Infoblox DNS appliances"],
         ["Rubrik backup appliances"], ["Cisco IOS switches"]],
        sheet_name="Assets",
    )
    parsed = parse_environment_file(content, "xlsx")
    env = parsed["environment"]
    assert set(env["platforms"]) == {"Linux", "Network Devices"}
    by_entry = {i["entry"]: i for i in parsed["interpretations"]}
    assert by_entry["Photon OS appliances"]["interpretation"] == "counted as platform Linux"
    assert by_entry["Infoblox DNS appliances"]["interpretation"] == "counted as platform Network Devices"
    assert by_entry["Rubrik backup appliances"]["interpretation"] == "counted as platform Linux"
    # ordering regression: "cisco ios" must still resolve to Network Devices,
    # never iOS (word-boundary + longest-match discipline unaffected by A10)
    assert by_entry["Cisco IOS switches"]["interpretation"] == "counted as platform Network Devices"


def test_a10_iot_and_mainframe_stay_unmapped():
    """ATT&CK v19.1 has no IoT/mainframe platform -- mapping these would be
    dishonest, so they must remain in the unmapped-assets assumption line.
    Pinned so a future session doesn't "fix" this by inventing a mapping."""
    content = _xlsx(
        [["Asset"], ["IOT Platform devices"], ["Mainframe z/OS billing platform"]],
        sheet_name="Assets",
    )
    parsed = parse_environment_file(content, "xlsx")
    assert parsed["environment"]["platforms"] == []
    by_entry = {i["entry"]: i for i in parsed["interpretations"]}
    assert by_entry["IOT Platform devices"]["interpretation"] == (
        "not recognized — ignored for platform filtering"
    )
    assert by_entry["Mainframe z/OS billing platform"]["interpretation"] == (
        "not recognized — ignored for platform filtering"
    )
    assert any("IOT Platform devices" in a for a in parsed["assumptions"])
    assert any("Mainframe z/OS billing platform" in a for a in parsed["assumptions"])


# --- Phase A6: severity/last-triggered use-case columns (new-column goldens) ---

def test_use_case_severity_and_last_triggered_columns_parsed():
    content = _xlsx([
        TEMPLATE_HEADERS + ["Severity", "Last Triggered"],
        ["Old critical rule", "T1059.001", "q", "", "Sysmon", "Enabled", "Critical", "never"],
        ["Fresh medium rule", "T1021.001", "q", "", "Sysmon", "Enabled", "Medium", "2026-07-01"],
        ["No health data", "T1047", "q", "", "Sysmon", "Enabled", "", ""],
    ])
    parsed = parse_use_case_file(content, "xlsx")
    assert "severity" in parsed["columns"] and "last_triggered" in parsed["columns"]
    r1, r2, r3 = parsed["rows"]
    assert r1["severity"] == "Critical" and r1["last_triggered"] == "never"
    assert r2["severity"] == "Medium" and r2["last_triggered"] == "2026-07-01"
    assert r3["severity"] is None and r3["last_triggered"] is None


def test_use_case_last_triggered_native_excel_date_normalized():
    import datetime as dt

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(TEMPLATE_HEADERS + ["Severity", "Last Triggered"])
    ws.append(["Rule with real date", "T1059.001", "q", "", "Sysmon", "Enabled", "High", dt.date(2026, 1, 15)])
    buf = io.BytesIO()
    wb.save(buf)
    parsed = parse_use_case_file(buf.getvalue(), "xlsx")
    assert parsed["rows"][0]["last_triggered"] == "2026-01-15"


# --- Phase A6: Log Sources health columns (Parser/Format, Normalized, Last Event Seen) ---

def test_log_sources_health_columns_parsed_and_old_layout_unaffected():
    content = _xlsx(
        [["Platform"], ["Windows"]],
        sheet_name="Assets",
        extra_sheets=[(
            "Log Sources",
            [
                ["Source", "Present?", "Parser / Format", "Normalized (Y/N)", "Last Event Seen"],
                ["Sysmon", "Yes", "Sysmon XML", "Yes", "2026-08-01"],
                ["Legacy Syslog", "Yes", "Raw syslog", "No", "2025-01-01"],
                ["CloudTrail", "Yes", "", "", ""],
            ],
        )],
    )
    parsed = parse_environment_file(content, "xlsx")
    assert parsed["log_sources"] == ["Sysmon", "Legacy Syslog", "CloudTrail"]
    health = parsed["log_source_health"]
    assert health["Sysmon"] == {
        "parser_format": "Sysmon XML", "normalized": True, "last_event_seen": "2026-08-01",
    }
    assert health["Legacy Syslog"] == {
        "parser_format": "Raw syslog", "normalized": False, "last_event_seen": "2025-01-01",
    }
    assert "CloudTrail" not in health  # no health columns populated -> no entry
    by_entry = {i["entry"]: i for i in parsed["interpretations"] if i["sheet"] == "Log Sources"}
    assert "normalized: No" in by_entry["Legacy Syslog"]["interpretation"]


def test_log_sources_old_two_column_layout_still_parses_identically():
    content = _xlsx(
        [["Platform"], ["Windows"]],
        sheet_name="Assets",
        extra_sheets=[("Log Sources", [["Source"], ["Sysmon"], ["CloudTrail"]])],
    )
    parsed = parse_environment_file(content, "xlsx")
    assert parsed["log_sources"] == ["Sysmon", "CloudTrail"]
    assert parsed["log_source_health"] == {}


# --- Phase A6: Assets CMDB header-synonym widening (ServiceNow/Lansweeper) ---

def test_assets_cmdb_header_synonyms_recognized_as_header_row():
    content = _xlsx(
        [["OS"], ["Windows Server 2022"], ["RHEL 9"]],
        sheet_name="Assets",
    )
    parsed = parse_environment_file(content, "xlsx")
    # "OS" header row skipped -> only the two real entries counted
    assert set(parsed["environment"]["platforms"]) == {"Windows", "Linux"}
    by_entry = {i["entry"] for i in parsed["interpretations"]}
    assert "OS" not in by_entry

    content = _xlsx(
        [["CI Type"], ["Windows Server 2022"]],
        sheet_name="Assets",
    )
    parsed = parse_environment_file(content, "xlsx")
    assert parsed["environment"]["platforms"] == ["Windows"]


# --- Phase A6 acceptance: the real shipped templates round-trip through ingest ---

def test_real_use_case_template_round_trips():
    content = (TEMPLATES_DIR / "scopewise-mitre-use-cases.xlsx").read_bytes()
    parsed = parse_use_case_file(content, "xlsx")
    assert parsed["row_count"] >= 1
    assert {"name", "tags", "logic", "description", "log_source", "enabled",
            "severity", "last_triggered"} <= set(parsed["columns"])
    assert not parsed["warnings"]


def test_real_environment_template_round_trips():
    content = (TEMPLATES_DIR / "scopewise-mitre-environment.xlsx").read_bytes()
    parsed = parse_environment_file(content, "xlsx")
    assert set(parsed["sheets_found"]) == {"assets", "log_sources", "tooling", "crown_jewels"}
    assert parsed["log_sources"]
    assert parsed["log_source_health"]  # the shipped example rows populate it
    assert parsed["environment"]["inventory_provided"] is True
    # "Read Me" is an unrecognized sheet name -> tolerated, never surfaced
    # as a data sheet or an unmapped-entry assumption.
    assert not any("Read Me" in a for a in parsed["assumptions"])
