"""Document parsing for DOCX and PDF files."""

import logging
import re
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


class DocumentType(str, Enum):
    """Document types recognized by parser."""

    SOW = "SOW"
    PROPOSAL = "Proposal"
    RFP = "RFP"
    PROJECT_PLAN = "ProjectPlan"
    HLD = "HLD"
    LLD = "LLD"
    SOP = "SOP"
    SLA = "SLA"
    SECURITY = "Security"
    OTHER = "Other"


@dataclass
class ParsedSection:
    """A section extracted from document."""

    heading: str
    level: int  # 1=H1, 2=H2, etc.
    content: str
    page_number: Optional[int] = None
    start_offset: int = 0
    end_offset: int = 0


@dataclass
class ParseResult:
    """Result from parsing a document."""

    status: str  # success | partial | failed
    raw_text: str
    sections: list[ParsedSection]
    page_count: int
    detected_type: Optional[DocumentType] = None
    language: str = "en"
    error_message: Optional[str] = None
    tokens_estimated: int = 0


class DocxParser:
    """Parse DOCX files."""

    @staticmethod
    async def parse(file_content: bytes) -> ParseResult:
        """Parse DOCX file."""
        try:
            from io import BytesIO

            from docx import Document
            from docx.enum.style import WD_STYLE_TYPE
            from docx.table import Table
            from docx.text.paragraph import Paragraph

            doc = Document(BytesIO(file_content))
            raw_text = []
            sections = []
            # DOCX has no intrinsic pages (layout is renderer-dependent),
            # but Word stamps <w:lastRenderedPageBreak/> where each page
            # began at last save, and explicit breaks are <w:br
            # w:type="page"/>. Counting both gives real page numbers for
            # Word-authored files; programmatically generated files with no
            # markers fall back to a chars-per-page estimate below.
            current_page = 1
            saw_page_breaks = False
            char_offset = 0

            # iter_inner_content() walks paragraphs AND tables in document
            # order -- many real-world SOW/RFP templates (not just this
            # project's TemplateLab samples) lay out their actual content
            # in tables rather than plain paragraphs. The old paragraphs-only
            # loop silently returned status="success" with raw_text="" for
            # every such document -- a document with real content parsing
            # to empty text, with no error surfaced anywhere.
            for block in doc.iter_inner_content():
                if isinstance(block, Table):
                    table_text = DocxParser._extract_table_text(block)
                    if not table_text:
                        continue
                    raw_text.append(table_text)
                    char_offset += len(table_text) + 1
                    if sections:
                        sections[-1].content += table_text + "\n"
                    continue

                if not isinstance(block, Paragraph):
                    continue

                breaks = block._p.xml.count("lastRenderedPageBreak") + block._p.xml.count(
                    'type="page"'
                )
                if breaks:
                    current_page += breaks
                    saw_page_breaks = True

                text = block.text.strip()
                if not text:
                    continue

                raw_text.append(text)
                char_offset += len(text) + 1

                # Detect heading level
                level = 0
                if block.style and block.style.name.startswith("Heading"):
                    # Extract number from "Heading 1", "Heading 2", etc.
                    try:
                        level = int(block.style.name.split()[-1])
                    except (IndexError, ValueError):
                        level = 0

                # If it's a heading, create a section
                if level > 0:
                    sections.append(
                        ParsedSection(
                            heading=text,
                            level=level,
                            content="",  # Will be filled as we parse
                            page_number=current_page,
                            start_offset=char_offset - len(text) - 1,
                        )
                    )
                elif sections:
                    # Add to last section's content
                    sections[-1].content += text + "\n"

            # Combine text
            full_text = "\n".join(raw_text)

            if saw_page_breaks:
                page_count = current_page
            else:
                # ponytail: ~3000 chars/page estimate; real layout needs a
                # renderer. Only used for files with no Word page markers.
                _CHARS_PER_PAGE = 3000
                page_count = max(1, -(-len(full_text) // _CHARS_PER_PAGE))
                for s in sections:
                    s.page_number = min(page_count, s.start_offset // _CHARS_PER_PAGE + 1)

            # Estimate tokens (rough: 1 token ≈ 4 chars)
            tokens = len(full_text) // 4

            # Detect document type
            detected_type = DocxParser._detect_type(full_text)

            return ParseResult(
                status="success",
                raw_text=full_text,
                sections=sections,
                page_count=page_count,
                detected_type=detected_type,
                tokens_estimated=tokens,
            )

        except Exception as e:
            logger.error(f"DOCX parsing error: {e}")
            return ParseResult(
                status="failed",
                raw_text="",
                sections=[],
                page_count=0,
                error_message=str(e),
            )

    @staticmethod
    def _extract_table_text(table) -> str:
        """Flattens a docx table's cell text into lines. Merged cells share
        the same underlying XML element across every grid position they
        span, so a naive per-cell loop would repeat that cell's text once
        per spanned column/row -- dedup by the underlying element's
        identity within each row to avoid that."""
        lines = []
        for row in table.rows:
            seen = set()
            cell_texts = []
            for cell in row.cells:
                cell_id = id(cell._tc)
                if cell_id in seen:
                    continue
                seen.add(cell_id)
                text = cell.text.strip()
                if text:
                    cell_texts.append(text)
            if cell_texts:
                lines.append(" | ".join(cell_texts))
        return "\n".join(lines)

    @staticmethod
    def _detect_type(text: str) -> DocumentType:
        """Detect document type from content."""
        text_lower = text.lower()

        # Simple pattern matching (can be improved with ML later).
        # RFP is checked BEFORE the bare "statement of work"/"sow" match --
        # RFPs routinely mention "Statement of Work" as boilerplate (the
        # deliverable the *awarded* vendor will produce, e.g. "the selected
        # vendor shall submit a Statement of Work within 10 business days
        # of award"), which would otherwise misclassify the RFP as a SOW
        # before this check ever ran, silently routing it through the
        # wrong ReviewAgent branch and rule set.
        if "request for proposal" in text_lower or re.search(r"\brfp\b", text_lower):
            return DocumentType.RFP
        elif "statement of work" in text_lower or "sow" in text_lower:
            return DocumentType.SOW
        elif "proposal" in text_lower:
            return DocumentType.PROPOSAL
        elif "project plan" in text_lower:
            return DocumentType.PROJECT_PLAN
        elif "high level design" in text_lower or "hld" in text_lower:
            return DocumentType.HLD
        elif "low level design" in text_lower or "lld" in text_lower:
            return DocumentType.LLD
        elif "standard operating procedure" in text_lower or "sop" in text_lower:
            return DocumentType.SOP
        elif "service level agreement" in text_lower or "sla" in text_lower:
            return DocumentType.SLA
        elif "security" in text_lower:
            return DocumentType.SECURITY

        return DocumentType.OTHER


class DocParser:
    """Parse legacy binary .doc files via the `antiword` CLI.

    # ponytail: shells out to antiword rather than a pure-Python OLE parser
    # -- legacy .doc's binary format has no good pure-Python extractor;
    # antiword is a small, purpose-built tool (not a full LibreOffice/
    # pandoc conversion chain). No section/heading detection (antiword
    # only gives flat text) -- upgrade to a heading-aware extractor if
    # section-level review accuracy on .doc files becomes a problem.
    """

    @staticmethod
    async def parse(file_content: bytes) -> ParseResult:
        import asyncio
        import subprocess
        import tempfile

        try:
            with tempfile.NamedTemporaryFile(suffix=".doc", delete=False) as tmp:
                tmp.write(file_content)
                tmp_path = tmp.name

            try:
                proc = await asyncio.create_subprocess_exec(
                    "antiword",
                    tmp_path,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                )
                stdout, stderr = await proc.communicate()
            finally:
                Path(tmp_path).unlink(missing_ok=True)

            if proc.returncode != 0:
                raise RuntimeError(stderr.decode("utf-8", errors="replace") or "antiword failed")

            full_text = stdout.decode("utf-8", errors="replace")
            tokens = len(full_text) // 4
            detected_type = DocxParser._detect_type(full_text)

            return ParseResult(
                status="success" if full_text.strip() else "partial",
                raw_text=full_text,
                sections=[],
                page_count=1,
                detected_type=detected_type,
                tokens_estimated=tokens,
            )

        except FileNotFoundError:
            logger.error("DOC parsing error: antiword is not installed")
            return ParseResult(
                status="failed",
                raw_text="",
                sections=[],
                page_count=0,
                error_message="Legacy .doc parsing requires antiword, which isn't installed",
            )
        except Exception as e:
            logger.error(f"DOC parsing error: {e}")
            return ParseResult(
                status="failed",
                raw_text="",
                sections=[],
                page_count=0,
                error_message=str(e),
            )


class ExcelParser:
    """Parse .xlsx files -- each sheet becomes a section, cell values
    joined row by row into that section's content."""

    @staticmethod
    async def parse(file_content: bytes) -> ParseResult:
        try:
            from io import BytesIO

            from openpyxl import load_workbook

            workbook = load_workbook(BytesIO(file_content), data_only=True, read_only=True)
            raw_text = []
            sections = []

            for sheet in workbook.worksheets:
                sheet_lines = []
                for row in sheet.iter_rows(values_only=True):
                    cells = [str(v) for v in row if v is not None]
                    if cells:
                        sheet_lines.append("\t".join(cells))

                sheet_text = "\n".join(sheet_lines)
                raw_text.append(f"--- Sheet: {sheet.title} ---\n{sheet_text}")
                sections.append(ParsedSection(heading=sheet.title, level=1, content=sheet_text))

            full_text = "\n".join(raw_text)
            tokens = len(full_text) // 4
            detected_type = DocxParser._detect_type(full_text)

            return ParseResult(
                status="success" if full_text.strip() else "partial",
                raw_text=full_text,
                sections=sections,
                page_count=len(workbook.worksheets),
                detected_type=detected_type,
                tokens_estimated=tokens,
            )

        except Exception as e:
            logger.error(f"Excel parsing error: {e}")
            return ParseResult(
                status="failed",
                raw_text="",
                sections=[],
                page_count=0,
                error_message=str(e),
            )


class CsvParser:
    """Parse .csv files via the stdlib csv module -- rows joined with tabs,
    no section detection (flat tabular data has no headings)."""

    @staticmethod
    async def parse(file_content: bytes) -> ParseResult:
        try:
            import csv
            from io import StringIO

            text = file_content.decode("utf-8-sig", errors="replace")
            reader = csv.reader(StringIO(text))
            lines = ["\t".join(row) for row in reader if row]
            full_text = "\n".join(lines)
            tokens = len(full_text) // 4
            detected_type = DocxParser._detect_type(full_text)

            return ParseResult(
                status="success" if full_text.strip() else "partial",
                raw_text=full_text,
                sections=[],
                page_count=1,
                detected_type=detected_type,
                tokens_estimated=tokens,
            )
        except Exception as e:
            logger.error(f"CSV parsing error: {e}")
            return ParseResult(
                status="failed",
                raw_text="",
                sections=[],
                page_count=0,
                error_message=str(e),
            )


class XlsParser:
    """Parse legacy binary .xls files via xlrd (openpyxl only reads the
    modern .xlsx format). Same sheet-as-section shape as ExcelParser."""

    @staticmethod
    async def parse(file_content: bytes) -> ParseResult:
        try:
            from io import BytesIO

            import xlrd

            workbook = xlrd.open_workbook(file_contents=file_content)
            raw_text = []
            sections = []

            for sheet in workbook.sheets():
                sheet_lines = []
                for row_idx in range(sheet.nrows):
                    cells = [str(v) for v in sheet.row_values(row_idx) if v not in (None, "")]
                    if cells:
                        sheet_lines.append("\t".join(cells))

                sheet_text = "\n".join(sheet_lines)
                raw_text.append(f"--- Sheet: {sheet.name} ---\n{sheet_text}")
                sections.append(ParsedSection(heading=sheet.name, level=1, content=sheet_text))

            full_text = "\n".join(raw_text)
            tokens = len(full_text) // 4
            detected_type = DocxParser._detect_type(full_text)

            return ParseResult(
                status="success" if full_text.strip() else "partial",
                raw_text=full_text,
                sections=sections,
                page_count=workbook.nsheets,
                detected_type=detected_type,
                tokens_estimated=tokens,
            )
        except Exception as e:
            logger.error(f"XLS parsing error: {e}")
            return ParseResult(
                status="failed",
                raw_text="",
                sections=[],
                page_count=0,
                error_message=str(e),
            )


class PdfParser:
    """Parse PDF files."""

    @staticmethod
    async def parse(file_content: bytes) -> ParseResult:
        """Parse PDF file."""
        try:
            from pypdf import PdfReader
            from io import BytesIO

            pdf = PdfReader(BytesIO(file_content))
            page_count = len(pdf.pages)
            raw_text = []
            sections = []

            page_texts = [page.extract_text() or "" for page in pdf.pages]

            # Image-only/scanned PDF: no text layer to extract. Fall back to
            # OCR (tesseract via pypdfium2 rendering) when available.
            ocr_used = False
            if page_count and sum(len(t.strip()) for t in page_texts) < 200:
                import asyncio

                ocr_texts = await asyncio.to_thread(PdfParser._ocr_pages, file_content)
                if ocr_texts is not None:
                    page_texts = ocr_texts + [""] * (page_count - len(ocr_texts))
                    ocr_used = True

            # Extract text from all pages
            for page_num, text in enumerate(page_texts, 1):
                raw_text.append(f"--- Page {page_num} ---\n{text}")

                # Try to detect sections by font size changes (simplified)
                # In a real implementation, use advanced PDF analysis
                lines = text.split("\n")
                for line in lines:
                    line = line.strip()
                    if len(line) > 5 and len(line) < 100:  # Heuristic for headings
                        if line.isupper() or (line[0].isupper() and len(line.split()) < 10):
                            sections.append(
                                ParsedSection(
                                    heading=line,
                                    level=1,
                                    content="",
                                    page_number=page_num,
                                )
                            )

            full_text = "\n".join(raw_text)
            tokens = len(full_text) // 4

            # "--- Page N ---" markers alone are not content: without this
            # check a scanned PDF parsed as "success" with an empty body and
            # produced a page of bogus missing-section findings downstream.
            body_chars = sum(len(t.strip()) for t in page_texts)
            if body_chars < 200:
                return ParseResult(
                    status="failed",
                    raw_text="",
                    sections=[],
                    page_count=page_count,
                    error_message=(
                        "No extractable text (scanned/image-only PDF?); "
                        + ("OCR found no text" if ocr_used else "OCR not available")
                    ),
                )

            # Detect document type
            detected_type = PdfParser._detect_type(full_text)

            return ParseResult(
                status="success",
                raw_text=full_text,
                sections=sections,
                page_count=page_count,
                detected_type=detected_type,
                tokens_estimated=tokens,
            )

        except Exception as e:
            logger.error(f"PDF parsing error: {e}")
            return ParseResult(
                status="failed",
                raw_text="",
                sections=[],
                page_count=0,
                error_message=str(e),
            )

    # ponytail: 30-page OCR cap at ~200 DPI -- a scanned contract beyond that
    # is rare, and OCR is CPU-bound on a shared VPS; raise if it ever bites.
    _OCR_MAX_PAGES = 30

    @staticmethod
    def _ocr_pages(file_content: bytes) -> Optional[list[str]]:
        """OCR each page; None when tesseract/pytesseract isn't installed
        (callers keep the no-text-layer failure path)."""
        import shutil

        if not shutil.which("tesseract"):
            logger.warning("Scanned PDF but tesseract binary not installed -- skipping OCR")
            return None
        try:
            import pypdfium2 as pdfium
            import pytesseract
        except ImportError:
            logger.warning("Scanned PDF but pypdfium2/pytesseract not installed -- skipping OCR")
            return None

        try:
            doc = pdfium.PdfDocument(file_content)
            texts = []
            for i in range(min(len(doc), PdfParser._OCR_MAX_PAGES)):
                image = doc[i].render(scale=200 / 72).to_pil()
                texts.append(pytesseract.image_to_string(image))
            logger.info(f"OCR extracted {sum(len(t) for t in texts)} chars from {len(texts)} pages")
            return texts
        except Exception as e:
            logger.error(f"OCR failed: {e}")
            return None

    @staticmethod
    def _detect_type(text: str) -> DocumentType:
        """Detect document type from content."""
        text_lower = text.lower()

        # RFP checked before the bare "statement of work"/"sow" match -- see
        # DocxParser._detect_type's comment above for why (RFP boilerplate
        # commonly mentions "Statement of Work" as the future deliverable).
        if "request for proposal" in text_lower or re.search(r"\brfp\b", text_lower):
            return DocumentType.RFP
        elif "statement of work" in text_lower or "sow" in text_lower:
            return DocumentType.SOW
        elif "proposal" in text_lower:
            return DocumentType.PROPOSAL
        elif "project plan" in text_lower:
            return DocumentType.PROJECT_PLAN
        elif "high level design" in text_lower or "hld" in text_lower:
            return DocumentType.HLD
        elif "low level design" in text_lower or "lld" in text_lower:
            return DocumentType.LLD
        elif "standard operating procedure" in text_lower or "sop" in text_lower:
            return DocumentType.SOP
        elif "service level agreement" in text_lower or "sla" in text_lower:
            return DocumentType.SLA
        elif "security" in text_lower:
            return DocumentType.SECURITY

        return DocumentType.OTHER


async def parse_document(
    file_content: bytes, file_type: str
) -> ParseResult:
    """
    Parse document based on file type.

    T-313-T-320: Main parsing pipeline
    """
    file_type = file_type.lower()

    if file_type in ("docx", "application/vnd.openxmlformats-officedocument.wordprocessingml.document"):
        return await DocxParser.parse(file_content)

    elif file_type in ("pdf", "application/pdf"):
        return await PdfParser.parse(file_content)

    elif file_type in ("doc", "application/msword"):
        return await DocParser.parse(file_content)

    elif file_type in (
        "xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ):
        return await ExcelParser.parse(file_content)

    elif file_type in ("xls", "application/vnd.ms-excel"):
        return await XlsParser.parse(file_content)

    elif file_type in ("csv", "text/csv"):
        return await CsvParser.parse(file_content)

    else:
        return ParseResult(
            status="failed",
            raw_text="",
            sections=[],
            page_count=0,
            error_message=f"Unsupported file type: {file_type}",
        )
