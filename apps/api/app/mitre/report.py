"""MITRE assessment reports: exec+detailed HTML/PDF and the XLSX gap
register (plan §10).

Numbers come ONLY from the stored summary/technique_results JSONB — never
recomputed here, never parsed out of narrative text. Every customer/LLM
string goes through _esc (HTML) or _guard (XLSX formula injection): rule
names, descriptions, exclusion reasons, and narrative output are all
attacker-controlled.

PDF rendering reuses the house WeasyPrint approach (app/scoring/report.py):
lazily imported because its native libs (Pango/Cairo) only exist in the
prod image — local dev fails soft at call time, not import time.
"""

import io
import logging
import os
from datetime import datetime, timezone

from app.scoring.report import _esc  # house escaper, stored-XSS lesson baked in

logger = logging.getLogger(__name__)

STATE_LABELS = {
    "covered": "Covered",
    "partial": "Partial",
    "not_covered": "Not covered",
    "not_applicable": "N/A",
}
STATE_COLORS = {
    "covered": "#10b981",
    "partial": "#f59e0b",
    "not_covered": "#f43f5e",
    "not_applicable": "#9ca3af",
}
FEASIBILITY_LABELS = {"short": "Short term (0–3 mo)", "mid": "Mid term (3–9 mo)", "long": "Long term (9–18 mo)"}
DOMAIN_LABELS = {"enterprise": "Enterprise", "ics": "ICS / OT", "mobile": "Mobile"}

# PDF appendix cap — a 5,000-row use-case appendix belongs in the XLSX, not
# a PDF. The cap is stated in the report when it bites.
MAX_APPENDIX_ROWS = 500

_NA_GROUPS = [
    ("Whole matrix not applicable", lambda r: "matrix:" in r),
    ("Platform not in the environment", lambda r: r.startswith("targets ")),
    ("Deprecated by MITRE", lambda r: r.startswith("deprecated in ATT&CK")),
    ("Customer-declared exclusions", lambda r: True),
]


def _pct_bar(pct: float, color: str = "#0057B8") -> str:
    width = max(0.0, min(100.0, float(pct or 0)))
    return (
        '<div style="background:#e5e7eb;border-radius:4px;height:10px;width:100%;">'
        f'<div style="background:{color};border-radius:4px;height:10px;width:{width}%;"></div></div>'
    )


def _state_chip(state: str) -> str:
    return (
        f'<span style="color:{STATE_COLORS.get(state, "#333")};font-weight:600;">'
        f"{_esc(STATE_LABELS.get(state, state))}</span>"
    )


def _traffic_color(pct) -> str:
    """Deterministic traffic-light band for the executive scorecard."""
    value = float(pct or 0)
    if value >= 50:
        return "#10b981"
    if value >= 15:
        return "#f59e0b"
    return "#f43f5e"


def _stacked_bar(covered, partial, not_covered, applicable) -> str:
    """CSS stacked horizontal bar for one tactic (deterministic widths)."""
    total = max(1, int(applicable or 0))
    seg = (
        "<span style='display:inline-block;height:9px;background:{c};width:{w}%;'></span>"
    )
    return (
        "<div style='background:#e5e7eb;border-radius:4px;height:9px;width:100%;"
        "font-size:0;line-height:0;overflow:hidden;'>"
        + seg.format(c="#10b981", w=round(100 * (covered or 0) / total, 2))
        + seg.format(c="#f59e0b", w=round(100 * (partial or 0) / total, 2))
        + seg.format(c="#f43f5e", w=round(100 * (not_covered or 0) / total, 2))
        + "</div>"
    )


def build_html_report(assessment, use_cases: list, compare=None, files=None) -> str:
    """Executive + detailed report as one self-contained HTML document
    (rebuilt in Phase 14e: cover → executive ≤2 pages → detailed →
    appendices, TOC with real page numbers, running header, deterministic
    HTML/CSS only — numbers come solely from the stored JSONB).

    assessment: MitreAssessment ORM row (or anything with the same
    attributes). use_cases: dicts {row_ref, name, description, log_source,
    enabled, mappings, mapping_status}. compare: optional
    service.compare_assessments() output (+ baseline_name) for the trend
    block. files: optional [{kind, filename, row_count}] for the cover.
    """
    from app.mitre import attack_data, plain_language

    summary = assessment.summary or {}
    params = assessment.params or {}
    overall = summary.get("overall", {})
    domains = summary.get("domains", {})
    gaps = summary.get("gaps", [])
    roadmap = summary.get("roadmap", {})
    narrative = summary.get("narrative", {})
    assumptions = summary.get("assumptions", [])
    not_applicable = summary.get("not_applicable", [])
    counts = summary.get("counts", {})
    index = attack_data.DEFAULT

    completed = assessment.completed_at.strftime("%Y-%m-%d %H:%M UTC") if assessment.completed_at else "—"
    intake = params.get("intake") or {}
    doc_title = intake.get("project_name") or assessment.name

    # Per-technique mapped rules (feeds why-phrases). Same inputs the drawer
    # explain endpoint uses — numbers/wording stay consistent across surfaces.
    mapped_by_technique: dict = {}
    for uc in use_cases:
        for m in uc.get("mappings") or []:
            mapped_by_technique.setdefault(m.get("technique_id"), []).append(
                {"name": uc.get("name"), "enabled": uc.get("enabled"),
                 "source": m.get("source"), "confidence": m.get("confidence")}
            )
    results = assessment.technique_results or []
    state_by_id = {r.get("technique_id"): r.get("state") for r in results}
    total_rules = counts.get("use_cases", len(use_cases))
    confidence_covered = (params.get("thresholds") or {}).get("confidence_covered", 0.7)
    gap_recs = narrative.get("gap_recommendations", {})
    ai_badge = (
        "<span class='badge ai'>AI-written text</span>"
        if narrative.get("generated_by") == "ai"
        else "<span class='badge'>Standard text</span>"
    )

    # --- cover ----------------------------------------------------------
    meta_rows = "".join(
        f"<tr><th>{_esc(label)}</th><td>{_esc(intake.get(key))}</td></tr>"
        for key, label in (
            ("project_name", "Organization / project"),
            ("scope_label", "Scope"),
            ("prepared_by", "Prepared by"),
            ("purpose_note", "Purpose"),
        )
        if intake.get(key)
    )
    files = files or []
    uc_file = next((f for f in files if f.get("kind") == "use_cases"), None)
    env_file = next((f for f in files if f.get("kind") == "environment"), None)
    env = params.get("environment") or {}
    env_lists = params.get("environment_lists") or {}
    upload_bits = [
        (f"Detection rules: {_esc(uc_file.get('filename'))} — " if uc_file else "Detection rules: ")
        + f"{_esc(total_rules)} rules ({_esc(counts.get('customer_tagged', 0))} tagged by you, "
        + f"{_esc(counts.get('keyword_tagged', 0))} keyword-matched, {_esc(counts.get('ai_tagged', 0))} AI-tagged, "
        + f"{_esc(counts.get('unmapped', 0))} unmapped, {_esc(counts.get('invalid', 0))} invalid)"
    ]
    if env_file or env.get("platforms"):
        upload_bits.append(
            (f"Environment: {_esc(env_file.get('filename'))} — " if env_file else "Environment: ")
            + f"platforms {_esc(', '.join(env.get('platforms') or []) or 'none detected')}"
            + (" · OT/ICS assets" if env.get("has_ics_assets") else "")
            + (" · managed mobile" if env.get("has_managed_mobile") else "")
            + f" · {len(env_lists.get('log_sources') or [])} log sources"
            + f" · {len(env_lists.get('tooling') or [])} tooling entries"
        )
    else:
        upload_bits.append(
            "Environment: none provided — full ATT&CK matrices assessed, the score is a lower bound"
        )
    upload_html = "".join(f"<p class='muted'>{b}</p>" for b in upload_bits)

    # --- executive: scorecard, top-5 fixes, roadmap glance, trend --------
    scorecard = "".join(
        "<div class='tile'>"
        f"<b style='color:{_traffic_color(d.get('strict_pct'))}'>{_esc(d.get('strict_pct'))}%</b>"
        f"{_esc(DOMAIN_LABELS.get(key, key))}<br>"
        f"<span class='muted'>{_esc(d.get('covered'))} of {_esc(d.get('applicable'))} techniques covered</span>"
        "</div>"
        for key, d in domains.items()
        if d.get("applicable", 0) > 0
    )
    gated_notes = "".join(
        f"<p class='muted'>{_esc(DOMAIN_LABELS.get(key, key))}: not assessed — "
        f"{_esc(next((n['reason'] for n in not_applicable if n.get('domain') == key), ''))}</p>"
        for key, d in domains.items()
        if d.get("applicable", 0) == 0
    )

    fixes_html = ""
    for g in gaps[:5]:
        tid = g.get("technique_id")
        described = plain_language.describe_technique(tid, index)
        relevance = g.get("threat_relevance") or []
        matters = (
            "Publicly reported technique of " + ", ".join(relevance)
            if relevance
            else ("Among the most-used attacker techniques in real intrusions"
                  if g.get("tier", 4) <= 2 else "A common supporting attacker behavior")
        )
        sketch = plain_language.detection_sketch(tid, g.get("via")) or g.get("hint") or ""
        fixes_html += (
            "<div class='fix'>"
            f"<p><strong>{_esc(tid)} {_esc(g.get('name'))}</strong> "
            f"<span class='badge'>{_esc(FEASIBILITY_LABELS.get(g.get('feasibility'), ''))}</span> "
            f"<a class='xref' href='#g-{_esc(tid)}'></a></p>"
            + (f"<p>{_esc(described.get('definition'))}</p>" if described.get("definition") else "")
            + f"<p class='muted'>Why it matters to you: {_esc(matters)}.</p>"
            + f"<p class='muted'>The fix: {_esc(sketch)}</p>"
            + "</div>"
        )

    short_items = roadmap.get("short", [])
    applicable_total = overall.get("applicable") or 0
    projected = (
        round(100 * ((overall.get("covered") or 0) + len(short_items)) / applicable_total, 1)
        if applicable_total
        else None
    )
    roadmap_glance = " · ".join(
        f"{_esc(FEASIBILITY_LABELS[b])}: {len(roadmap.get(b, []))}" for b in ("short", "mid", "long")
    )
    projection_html = (
        f"<p><strong>Effort to impact:</strong> completing just the short-term items "
        f"raises coverage from {_esc(overall.get('strict_pct'))}% to about {_esc(projected)}%.</p>"
        if projected is not None and short_items
        else ""
    )

    trend_html = ""
    if compare:
        delta = (compare.get("overall_delta") or {}).get("strict_pct")
        arrow = "▲" if (delta or 0) > 0 else ("▼" if (delta or 0) < 0 else "•")
        color = "#10b981" if (delta or 0) > 0 else ("#f43f5e" if (delta or 0) < 0 else "#6b7280")
        trend_html = (
            "<h3>Trend vs your previous run</h3>"
            f"<p><span style='color:{color};font-weight:bold'>{arrow} "
            f"{_esc(abs(delta) if delta is not None else 0)} points</span> vs "
            f"“{_esc(compare.get('baseline_name'))}”"
            f" — newly covered: {len(compare.get('newly_covered') or [])}, "
            f"regressed: {len(compare.get('regressed') or [])}, "
            f"applicability changed: {len(compare.get('na_changed') or [])}."
            + (" <span class='muted'>ATT&CK versions differ between runs — "
               "version-drift techniques were skipped.</span>"
               if compare.get("attack_version_mismatch") else "")
            + "</p>"
        )

    # --- detailed: stacked tactic bars + one-liners ----------------------
    tactic_sections = ""
    for key, d in domains.items():
        if d.get("applicable", 0) == 0:
            continue
        rows = ""
        for t in d.get("tactics", []):
            line = plain_language.TACTIC_LINES.get(t.get("shortname"))
            rows += (
                "<tr><td style='width:22%'>"
                f"<strong>{_esc(t.get('name'))}</strong>"
                + (f"<br><span class='muted'>{_esc(line)}</span>" if line else "")
                + "</td>"
                f"<td class='num'>{_esc(t.get('covered'))}/{_esc(t.get('applicable'))}"
                f"<br><span class='muted'>{_esc(t.get('strict_pct'))}%</span></td>"
                f"<td style='width:45%'>{_stacked_bar(t.get('covered'), t.get('partial'), t.get('not_covered'), t.get('applicable'))}</td></tr>"
            )
        tactic_sections += (
            f"<h3>{_esc(DOMAIN_LABELS.get(key, key))} — coverage by attack stage</h3>"
            f"<table><tbody>{rows}</tbody></table>"
        )

    # --- detailed: mini parent-level heatmap grids -----------------------
    heatmap_sections = ""
    for key, d in domains.items():
        if d.get("applicable", 0) == 0:
            continue
        cells = "".join(
            f"<span class='cell' style='background:{STATE_COLORS.get(r.get('state'), '#9ca3af')}' "
            f"title='{_esc(r.get('technique_id'))}'>{_esc(r.get('technique_id'))}</span>"
            for r in results
            if r.get("domain") == key and "." not in (r.get("technique_id") or "")
        )
        heatmap_sections += (
            f"<h3>{_esc(DOMAIN_LABELS.get(key, key))} — technique map (parent level)</h3>"
            f"<div class='heatmap'>{cells}</div>"
        )

    # --- detailed: gap register grouped by feasibility -------------------
    gap_sections = ""
    for bucket in ("short", "mid", "long"):
        bucket_gaps = [g for g in gaps if g.get("feasibility") == bucket]
        if not bucket_gaps:
            continue
        entries = ""
        for g in bucket_gaps:
            tid = g.get("technique_id")
            result = next((r for r in results if r.get("technique_id") == tid), None) or {
                "technique_id": tid, "state": g.get("state"), "na_reason": None,
            }
            mapped = mapped_by_technique.get(tid, [])
            why = plain_language.derive_why(
                result, mapped, total_rules=total_rules,
                sub_states=plain_language.sub_states_for(result, mapped, state_by_id, index),
                confidence_covered=confidence_covered,
            )
            sketch = plain_language.detection_sketch(tid, g.get("via"))
            via_line = (
                f"Uses logs you already collect: {g.get('via')}" if bucket == "short" and g.get("via")
                else f"Onboard telemetry from tooling you own: {g.get('via')}" if bucket == "mid" and g.get("via")
                else "Needs a new telemetry capability"
            )
            tier = g.get("tier", 4)
            relevance = g.get("threat_relevance") or []
            entries += (
                f"<div class='gap' id='g-{_esc(tid)}'>"
                f"<p><span class='num muted'>#{_esc(g.get('rank'))}</span> "
                f"<strong>{_esc(tid)} {_esc(g.get('name'))}</strong> "
                f"<span class='badge'>{'P' + str(tier) if tier < 4 else 'Unranked'}</span> "
                + ("<span class='badge threat'>Threat match: " + _esc(", ".join(relevance)) + "</span> "
                   if relevance else "")
                + f"{_state_chip(g.get('state', ''))}</p>"
                f"<p><em>Why it's a gap:</em> {_esc(why)}</p>"
                + (f"<p><em>What good looks like:</em> {_esc(sketch)}</p>" if sketch else "")
                + f"<p class='muted'>{_esc(via_line)}</p>"
                + (f"<p>{ai_badge} {_esc(gap_recs.get(tid))}</p>" if gap_recs.get(tid)
                   else f"<p class='muted'>{_esc(g.get('hint') or '')}</p>")
                + "</div>"
            )
        gap_sections += (
            f"<h3 class='bucket {bucket}'>{_esc(FEASIBILITY_LABELS[bucket])} — "
            f"{len(bucket_gaps)} item{'' if len(bucket_gaps) == 1 else 's'}</h3>"
            f"<p class='muted'>{_esc(narrative.get('roadmap_prose', {}).get(bucket, ''))}</p>"
            + entries
        )

    # --- appendices ------------------------------------------------------
    register_rows = "".join(
        f"<tr><td>{_esc(r.get('technique_id'))}</td>"
        f"<td>{_esc((index.get(r.get('technique_id')) or {}).get('name', ''))}</td>"
        f"<td>{_esc(DOMAIN_LABELS.get(r.get('domain'), r.get('domain')))}</td>"
        f"<td>{_state_chip(r.get('state', ''))}</td>"
        f"<td class='num'>{len(r.get('use_case_refs', []))}</td></tr>"
        for r in results
    )
    assumptions_html = "".join(f"<li>{_esc(a)}</li>" for a in assumptions)
    na_sections = ""
    remaining = list(not_applicable)
    for title, match in _NA_GROUPS:
        grouped = [n for n in remaining if match(n.get("reason") or "")]
        remaining = [n for n in remaining if n not in grouped]
        if not grouped:
            continue
        rows = "".join(
            f"<tr><td>{_esc(n.get('technique_id'))}</td>"
            f"<td>{_esc(DOMAIN_LABELS.get(n.get('domain'), n.get('domain')))}</td>"
            f"<td>{_esc(n.get('reason'))}</td></tr>"
            for n in grouped
        )
        na_sections += (
            f"<h3>{_esc(title)} ({len(grouped)})</h3>"
            "<table><thead><tr><th>Technique</th><th>Matrix</th><th>Reason</th></tr></thead>"
            f"<tbody>{rows}</tbody></table>"
        )

    # Phase 14g: "How we read your files" — the parser's per-entry evidence.
    interpretations = env_lists.get("interpretations") or []
    how_read_html = ""
    if interpretations:
        rows = "".join(
            f"<tr><td>{_esc(i.get('entry'))}</td><td>{_esc(i.get('sheet'))}</td>"
            f"<td>{_esc(i.get('interpretation'))}</td></tr>"
            for i in interpretations
        )
        how_read_html = (
            "<h2 id='how-read'>Appendix: how we read your files</h2>"
            "<p class='muted'>Every environment entry and what the parser did with it — "
            "the evidence behind platform filtering and roadmap feasibility.</p>"
            "<table><thead><tr><th>Your entry</th><th>Sheet</th><th>How it was read</th></tr></thead>"
            f"<tbody>{rows}</tbody></table>"
        )

    appendix_note = ""
    shown_use_cases = sorted(use_cases, key=_row_ref_sort_key)
    if len(shown_use_cases) > MAX_APPENDIX_ROWS:
        shown_use_cases = shown_use_cases[:MAX_APPENDIX_ROWS]
        appendix_note = (
            f"<p class='muted'>Showing the first {MAX_APPENDIX_ROWS} of "
            f"{len(use_cases)} rules — the XLSX export contains all of them.</p>"
        )
    uc_rows = "".join(
        f"<tr><td>{_esc(uc.get('row_ref'))}</td><td>{_esc(uc.get('name'))}</td>"
        f"<td>{'Enabled' if uc.get('enabled') else ('Disabled' if uc.get('enabled') is False else 'Unknown')}</td>"
        f"<td>{_esc(', '.join(m.get('technique_id', '') for m in (uc.get('mappings') or [])) or '—')}</td>"
        f"<td>{_esc(_MAPPING_STATUS_PLAIN_XLSX.get(uc.get('mapping_status'), uc.get('mapping_status')))}</td>"
        f"<td>{_esc(uc.get('log_source') or '')}</td></tr>"
        for uc in shown_use_cases
    )

    # --- audit footer ----------------------------------------------------
    thresholds = params.get("thresholds", {})
    models_used = params.get("models_used", {})
    footer_bits = [
        f"ATT&amp;CK v{_esc(assessment.attack_version)}",
        f"generated {_esc(datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC'))}",
        f"run completed {_esc(completed)}",
        f"narrative: {_esc(narrative.get('generated_by', 'n/a'))}"
        + (f" ({_esc(narrative.get('model_used'))})" if narrative.get("model_used") else ""),
    ]
    if os.getenv("GIT_SHA"):
        footer_bits.append(f"build {_esc(os.getenv('GIT_SHA'))}")
    if models_used:
        footer_bits.append(
            "models: " + _esc("; ".join(f"{k}: {', '.join(v)}" for k, v in models_used.items()))
        )
    siem = params.get("siem") or {}
    if siem:
        # Phase 13d provenance: which SIEM, manual vs scheduled, when.
        # Non-secret fields only (workspace_ref never holds credentials).
        source_bits = [
            "source: Microsoft Sentinel pull",
            _esc(siem.get("trigger") or "manual"),
        ]
        if (siem.get("workspace_ref") or {}).get("workspace"):
            source_bits.append(_esc(siem["workspace_ref"]["workspace"]))
        if siem.get("connection_name"):
            source_bits.append(_esc(siem["connection_name"]))
        if siem.get("pulled_at"):
            source_bits.append("pulled " + _esc(str(siem["pulled_at"])[:16]))
        footer_bits.append(" · ".join(source_bits))
    if thresholds:
        footer_bits.append(
            _esc(
                f"thresholds: covered≥{thresholds.get('confidence_covered')}, "
                f"partial≥{thresholds.get('confidence_partial_floor')}, "
                f"partial credit {thresholds.get('partial_credit')}, "
                f"disabled counts: {'yes' if thresholds.get('count_disabled_as_coverage') else 'no'}"
            )
        )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>MITRE ATT&amp;CK Coverage Assessment — {_esc(assessment.name)}</title>
<style>
@page {{
  size: A4; margin: 1.6cm 1.4cm 1.8cm;
  @top-center {{ content: string(doctitle); font-size: 9px; color: #6b7280; }}
  @bottom-right {{ content: "Page " counter(page) " of " counter(pages); font-size: 9px; color: #6b7280; }}
}}
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: Arial, 'Liberation Sans', Helvetica, sans-serif; color: #333; line-height: 1.5; font-size: 12px; }}
.container {{ max-width: 900px; margin: 0 auto; padding: 8px; }}
h1 {{ font-size: 22px; color: #0057B8; margin-bottom: 2px; string-set: doctitle content(); }}
h2 {{ font-size: 15px; color: #003D82; margin: 18px 0 6px; border-bottom: 2px solid #0057B8; padding-bottom: 3px; }}
h3 {{ font-size: 13px; margin: 12px 0 4px; }}
h3.bucket {{ padding: 3px 6px; border-radius: 4px; }}
h3.short {{ background: #d1fae5; }}
h3.mid {{ background: #fef3c7; }}
h3.long {{ background: #e5e7eb; }}
p {{ margin: 4px 0; }}
.muted {{ color: #6b7280; font-size: 11px; }}
.headline {{ font-size: 40px; font-weight: bold; color: #0057B8; }}
.tiles {{ display: flex; gap: 8px; margin: 8px 0; }}
.tile {{ flex: 1; background: #f3f4f6; border-radius: 6px; padding: 8px; text-align: center; }}
.tile b {{ display: block; font-size: 18px; }}
.badge {{ display: inline-block; border: 1px solid #d1d5db; border-radius: 9px; padding: 0 6px; font-size: 9px; color: #6b7280; vertical-align: middle; }}
.badge.ai {{ background: #ede9fe; border-color: #ddd6fe; color: #6d28d9; }}
.badge.threat {{ background: #ede9fe; border-color: #ddd6fe; color: #6d28d9; }}
.fix {{ border: 1px solid #e5e7eb; border-radius: 6px; padding: 6px 8px; margin: 6px 0; page-break-inside: avoid; }}
.gap {{ border-bottom: 1px solid #e5e7eb; padding: 6px 0; page-break-inside: avoid; }}
.heatmap {{ line-height: 1.15; }}
.heatmap .cell {{ display: inline-block; color: #fff; font-size: 7px; padding: 1px 3px; margin: 1px; border-radius: 2px; }}
table {{ width: 100%; border-collapse: collapse; margin: 6px 0 10px; font-size: 11px; }}
th {{ text-align: left; background: #f3f4f6; padding: 4px 6px; border-bottom: 1px solid #d1d5db; }}
td {{ padding: 4px 6px; border-bottom: 1px solid #e5e7eb; vertical-align: top; }}
td.num {{ text-align: right; white-space: nowrap; }}
ul {{ margin: 4px 0 4px 18px; }}
.cover-meta th {{ background: none; width: 30%; color: #6b7280; font-weight: normal; }}
.toc {{ margin: 10px 0 0 0; }}
.toc a {{ text-decoration: none; color: #333; }}
.toc li {{ margin: 2px 0; }}
.toc a::after {{ content: " — p. " target-counter(attr(href), page); color: #6b7280; }}
a.xref {{ text-decoration: none; color: #0057B8; font-size: 10px; }}
a.xref::after {{ content: "details p. " target-counter(attr(href), page); }}
.page-break {{ page-break-before: always; }}
.footer {{ margin-top: 20px; padding-top: 8px; border-top: 1px solid #d1d5db; font-size: 10px; color: #6b7280; }}
</style>
</head>
<body><div class="container">

<!-- ============================= COVER ============================= -->
<h1>MITRE ATT&amp;CK Coverage Assessment</h1>
<p><strong>{_esc(doc_title)}</strong>{'' if doc_title == assessment.name else f" — {_esc(assessment.name)}"}</p>
<table class="cover-meta"><tbody>{meta_rows}
<tr><th>ATT&amp;CK version</th><td>v{_esc(assessment.attack_version)}</td></tr>
<tr><th>Run completed</th><td>{_esc(completed)}</td></tr>
</tbody></table>
<div class="tiles" style="margin-top:14px">
  <div class="tile"><b class="headline">{_esc(overall.get('strict_pct'))}%</b>coverage of applicable techniques<br>
  <span class="muted">weighted (partial = half): {_esc(overall.get('weighted_pct'))}%</span></div>
</div>
<p>Of the {_esc(overall.get('applicable'))} attacker techniques that apply to your
environment, your detection rules can catch {_esc(overall.get('covered'))}
(plus {_esc(overall.get('partial'))} partially). Early SIEM detection programs typically
start under 10% — the roadmap in this report matters more than the grade.</p>
<h3>What this assessment is based on</h3>
{upload_html}
<p class="muted">Methodology: coverage is computed deterministically against the pinned
MITRE ATT&amp;CK v{_esc(assessment.attack_version)} dataset; techniques impossible in this
environment (or excluded by you) leave the denominator as “not applicable”.
A technique counts as covered when at least one enabled rule maps to it with qualifying
confidence. This assessment scores detection <em>presence</em>, not rule efficacy.</p>
<h3>Contents</h3>
<ul class="toc">
<li><a href="#exec">Executive summary</a></li>
<li><a href="#tactics">Coverage by attack stage</a></li>
<li><a href="#gapreg">Gap register &amp; recommendations</a></li>
<li><a href="#register">Appendix: technique register</a></li>
<li><a href="#na">Appendix: not-applicable techniques</a></li>
<li><a href="#assumptions">Appendix: assumptions</a></li>
{'<li><a href="#how-read">Appendix: how we read your files</a></li>' if how_read_html else ''}
<li><a href="#mappings">Appendix: rule mappings</a></li>
</ul>

<!-- ====================== EXECUTIVE (max 2 pages) ================== -->
<h2 id="exec" class="page-break">Executive summary</h2>
<div class="tiles">{scorecard}</div>
{gated_notes}
<p>{_esc(narrative.get('executive_summary', ''))} {ai_badge}</p>
<h3>Top 5 things to fix first</h3>
{fixes_html or "<p class='muted'>No gaps — nothing to fix.</p>"}
<p><strong>Roadmap at a glance:</strong> {roadmap_glance}</p>
{projection_html}
{trend_html}

<!-- =========================== DETAILED ============================ -->
<h2 id="tactics" class="page-break">Coverage by attack stage</h2>
{tactic_sections}
{heatmap_sections}

<h2 id="gapreg">Gap register &amp; recommendations ({len(gaps)})</h2>
<p class="muted">Grouped by how soon you could realistically build each detection.
Each entry explains why it is a gap and what a good detection would watch for.</p>
{gap_sections}

<!-- ========================== APPENDICES =========================== -->
<h2 id="register" class="page-break">Appendix: technique register ({len(results)})</h2>
<table><thead><tr><th>Technique</th><th>Name</th><th>Matrix</th><th>State</th><th>Rules</th></tr></thead>
<tbody>{register_rows}</tbody></table>

<h2 id="na">Appendix: not-applicable techniques ({len(not_applicable)})</h2>
<p class="muted">These techniques leave the coverage denominator — the headline percentage
makes no claim about them.</p>
{na_sections}

<h2 id="assumptions">Appendix: assumptions</h2>
<ul>{assumptions_html or '<li>None.</li>'}</ul>

{how_read_html}

<h2 id="mappings">Appendix: rule mappings ({len(use_cases)})</h2>
{appendix_note}
<table><thead><tr><th>Row</th><th>Rule</th><th>Status</th><th>Techniques</th><th>How it was mapped</th><th>Log source</th></tr></thead>
<tbody>{uc_rows}</tbody></table>

<div class="footer">{' · '.join(footer_bits)}</div>
</div></body></html>"""


def generate_pdf(html_content: str) -> bytes:
    """HTML -> PDF. Lazy import: WeasyPrint's native libs exist only in the
    prod image (Dockerfile.prod) — see app/scoring/report.py precedent."""
    from weasyprint import HTML

    return HTML(string=html_content).write_pdf()


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------


def _guard(value):
    """Excel formula-injection guard (plan §10): a leading =, +, - or @ in an
    attacker-controlled string (rule names, descriptions, reasons) would
    execute as a formula when the register opens in Excel."""
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        return "'" + value
    return value


# Phase 14c: fill colors (ARGB-less hex; light fills, dark text stays legible)
_XLSX_STATE_FILLS = {
    "covered": "C6EFCE",
    "partial": "FFE699",
    "not_covered": "FFC7CE",
    "not_applicable": "D9D9D9",
}
_XLSX_TIER_FILLS = {1: "F8CBAD", 2: "FFE699", 3: "FFF2CC"}
_XLSX_FEAS_FILLS = {"short": "C6EFCE", "mid": "FFE699", "long": "D9D9D9"}
_STATE_PLAIN_XLSX = {
    "covered": "A rule detects this",
    "partial": "Half-covered",
    "not_covered": "No rule detects this",
    "not_applicable": "Doesn't apply to this environment",
}
_MAPPING_STATUS_PLAIN_XLSX = {
    "customer_tagged": "You tagged this",
    "keyword_tagged": "Matched by tool/technique keyword (no AI)",
    "ai_tagged": "AI-suggested — verify",
    "manual": "Edited by a reviewer",
    "unmapped": "Could not be mapped",
    "invalid": "Tags were invalid — treated as untagged",
}


def _row_ref_sort_key(uc: dict):
    """Numeric-aware sort for row refs ('Rules:2' before 'Rules:10')."""
    import re as _re

    ref = str(uc.get("row_ref") or "")
    return (
        _re.sub(r"\d+", "#", ref),
        [int(n) for n in _re.findall(r"\d+", ref)],
    )


def build_xlsx_export(assessment, use_cases: list) -> bytes:
    """The detailed gap register as a 9-sheet workbook (Phase 14c polish):
    'Read Me' guide sheet first, colored state/priority/feasibility cells,
    frozen headers + auto-filter + wrapped text everywhere, technique names
    + plain-words 'Why' column, numerically sorted rule rows. Computed
    numbers are untouched — styling and wording only."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from app.mitre import attack_data, plain_language

    summary = assessment.summary or {}
    params = assessment.params or {}
    overall = summary.get("overall", {})
    narrative = summary.get("narrative", {})
    gap_recs = narrative.get("gap_recommendations", {})
    index = attack_data.DEFAULT

    wb = Workbook()
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    def fill(color):
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def sheet(title, headers, rows, widths, first=False, filters=True):
        ws = wb.active if first else wb.create_sheet()
        ws.title = title
        ws.append([_guard(h) for h in headers])
        for cell in ws[1]:
            cell.font = bold
        for row in rows:
            ws.append([_guard(v) for v in row])
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap
        ws.freeze_panes = "A2"
        if filters and rows:
            ws.auto_filter.ref = ws.dimensions
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        return ws

    # ---------------------------------------------------------------- Read Me
    strict_pct = overall.get("strict_pct")
    readme_rows = [
        ["This workbook is the full detail behind your MITRE ATT&CK coverage assessment."],
        [],
        ["The three key numbers"],
        [f"Coverage: {strict_pct}%",
         f"{overall.get('covered')} of {overall.get('applicable')} techniques that "
         "apply to your environment have at least one detection rule."],
        [f"Gaps to work on: {len(summary.get('gaps', []))}",
         "Ranked by priority in the 'Gaps & Recommendations' sheet — start at the top."],
        [f"Rules analyzed: {(summary.get('counts') or {}).get('use_cases', len(use_cases))}",
         "Your uploaded detection rules — see 'Use-Case Mappings' for what each one maps to."],
        [f"Is {strict_pct}% bad? Probably not: early SIEM detection programs typically "
         "start under 10%, because ATT&CK counts every known attacker technique. "
         "The roadmap matters more than the grade."],
        [],
        ["What each sheet contains"],
        ["Summary", "The headline numbers with a plain-words explanation of each."],
        ["Coverage by Tactic", "Coverage per attack stage (tactic), per matrix."],
        ["Technique Register", "Every technique: its state, why, and the rules mapped to it."],
        ["Use-Case Mappings", "Your rules, one per row, with how each was mapped."],
        ["Gaps & Recommendations", "What's missing, grouped by how soon you could fix it."],
        ["Roadmap", "The same gaps as a short/mid/long-term work plan."],
        ["Not Applicable", "Techniques that don't count toward your score, with reasons."],
        ["Assumptions", "What we had to assume — read before trusting the numbers."],
        [],
        ["Color legend"],
        ["Covered", "A rule detects this technique."],
        ["Partial", "Only a disabled rule, a low-confidence mapping, or a sub-technique reaches it."],
        ["Not covered", "No rule detects it — this is a gap."],
        ["N/A", "Doesn't apply to your environment; excluded from the score."],
    ]
    ws_readme = sheet("Read Me", ["How to read this workbook", ""],
                      readme_rows, [46, 100], first=True, filters=False)
    for label, color in (("Covered", "covered"), ("Partial", "partial"),
                         ("Not covered", "not_covered"), ("N/A", "not_applicable")):
        for row in ws_readme.iter_rows(min_row=2, max_col=1):
            if row[0].value == label:
                row[0].fill = fill(_XLSX_STATE_FILLS[color])
    for row in ws_readme.iter_rows(min_row=2, max_col=1):
        if row[0].value in ("The three key numbers", "What each sheet contains", "Color legend"):
            row[0].font = bold

    # ---------------------------------------------------------------- Summary
    domain_rows = [
        [f"{DOMAIN_LABELS.get(k, k)} coverage %", d.get("strict_pct"),
         f"Coverage within the {DOMAIN_LABELS.get(k, k)} matrix only."]
        for k, d in summary.get("domains", {}).items()
    ]
    # Phase 14d: optional project metadata rows (only when provided).
    intake = params.get("intake") or {}
    metadata_rows = [
        [label, intake[key], meaning]
        for key, label, meaning in (
            ("project_name", "Organization / project", "Who this assessment is for."),
            ("scope_label", "Scope", "The department or scope this run covers."),
            ("prepared_by", "Prepared by", "Who prepared this assessment."),
            ("purpose_note", "Purpose", "Why this assessment was run."),
        )
        if intake.get(key)
    ]
    sheet(
        "Summary",
        ["Metric", "Value", "What it means"],
        [
            ["Assessment", assessment.name, "The name this run was saved under."],
        ]
        + metadata_rows
        + [
            ["ATT&CK version", assessment.attack_version,
             "The MITRE ATT&CK release the assessment is pinned to."],
            ["Completed", str(assessment.completed_at or ""), ""],
            ["Coverage %", overall.get("strict_pct"),
             "Of the techniques that apply to your environment, the share with at "
             "least one qualifying detection rule."],
            ["Weighted coverage %", overall.get("weighted_pct"),
             "Same, but half-covered techniques count as 0.5 instead of 0."],
            ["Covered", overall.get("covered"), "Techniques at least one enabled rule detects."],
            ["Partial", overall.get("partial"),
             "Techniques reached only by a disabled rule, a low-confidence mapping, "
             "or a covered sub-technique."],
            ["Not covered", overall.get("not_covered"), "Techniques no rule detects — the gaps."],
            ["Not applicable", overall.get("not_applicable"),
             "Techniques excluded from the score (wrong platform, excluded by you, or deprecated)."],
            ["Applicable techniques", overall.get("applicable"),
             "The denominator: techniques that apply to your environment."],
            ["Narrative", narrative.get("generated_by", ""),
             "Whether recommendation wording was AI-written or standard template text "
             "(numbers are computed either way)."],
            ["Thresholds", str(params.get("thresholds", {})),
             "The confidence cut-offs used to count a mapping as coverage."],
        ]
        + domain_rows,
        [26, 44, 78],
    )

    # ------------------------------------------------------ Coverage by Tactic
    sheet(
        "Coverage by Tactic",
        ["Matrix", "Tactic", "Covered", "Partial", "Not covered", "N/A", "Applicable", "Coverage %", "Weighted %"],
        [
            [DOMAIN_LABELS.get(dk, dk), t.get("name"), t.get("covered"), t.get("partial"),
             t.get("not_covered"), t.get("not_applicable"), t.get("applicable"),
             t.get("strict_pct"), t.get("weighted_pct")]
            for dk, d in summary.get("domains", {}).items()
            for t in d.get("tactics", [])
        ],
        [12, 26, 10, 10, 12, 8, 12, 12, 12],
    )

    # ------------------------------------------------------- Technique Register
    rules_by_technique: dict = {}
    mapped_by_technique: dict = {}
    for uc in use_cases:
        for m in uc.get("mappings") or []:
            tid = m.get("technique_id")
            rules_by_technique.setdefault(tid, []).append(uc.get("name", ""))
            mapped_by_technique.setdefault(tid, []).append(
                {"name": uc.get("name"), "enabled": uc.get("enabled"),
                 "source": m.get("source"), "confidence": m.get("confidence")}
            )
    results = assessment.technique_results or []
    state_by_id = {r.get("technique_id"): r.get("state") for r in results}
    total_rules = (summary.get("counts") or {}).get("use_cases", len(use_cases))
    tactic_names = {
        (domain, t["id"]): t["name"]
        for domain in index.domains
        for t in index.tactics(domain)
    }
    confidence_covered = (params.get("thresholds") or {}).get("confidence_covered", 0.7)

    def register_row(r):
        tid = r.get("technique_id")
        mapped = mapped_by_technique.get(tid, [])
        why = plain_language.derive_why(
            r, mapped,
            total_rules=total_rules,
            sub_states=plain_language.sub_states_for(r, mapped, state_by_id, index),
            confidence_covered=confidence_covered,
        )
        return [
            tid,
            (index.get(tid) or {}).get("name", ""),
            DOMAIN_LABELS.get(r.get("domain"), r.get("domain")),
            ", ".join(tactic_names.get((r.get("domain"), t), t) for t in r.get("tactics", [])),
            STATE_LABELS.get(r.get("state"), r.get("state")),
            _STATE_PLAIN_XLSX.get(r.get("state"), ""),
            why,
            len(r.get("use_case_refs", [])),
            "; ".join(rules_by_technique.get(tid, [])),
        ]

    ws_reg = sheet(
        "Technique Register",
        ["Technique", "Name", "Matrix", "Tactics", "State", "In plain words", "Why", "Mapped rules", "Rule names"],
        [register_row(r) for r in results],
        [12, 30, 10, 26, 12, 26, 70, 12, 50],
    )
    for i, r in enumerate(results, start=2):
        color = _XLSX_STATE_FILLS.get(r.get("state"))
        if color:
            ws_reg.cell(row=i, column=5).fill = fill(color)

    # -------------------------------------------------------- Use-Case Mappings
    sorted_ucs = sorted(use_cases, key=_row_ref_sort_key)
    sheet(
        "Use-Case Mappings",
        ["Row", "Rule name", "Enabled", "How it was mapped", "Techniques", "Confidence", "Source", "Log source", "Description", "Logic"],
        [
            [uc.get("row_ref"), uc.get("name"),
             {True: "yes", False: "no"}.get(uc.get("enabled"), "unknown"),
             _MAPPING_STATUS_PLAIN_XLSX.get(uc.get("mapping_status"), uc.get("mapping_status")),
             ", ".join(m.get("technique_id", "") for m in (uc.get("mappings") or [])),
             ", ".join(str(m.get("confidence", "")) for m in (uc.get("mappings") or [])),
             ", ".join(m.get("source", "") for m in (uc.get("mappings") or [])),
             uc.get("log_source") or "", uc.get("description") or "",
             uc.get("logic") or ""]  # Phase 7; query text is attacker-controlled — _guard applies
            for uc in sorted_ucs
        ],
        [10, 42, 9, 30, 22, 12, 12, 16, 60, 60],
    )

    # --------------------------------------------- Gaps grouped by feasibility
    ws_gaps = wb.create_sheet()
    ws_gaps.title = "Gaps & Recommendations"
    gap_headers = ["Rank", "Technique", "Name", "Priority", "State", "Log source to use", "Recommendation"]
    ws_gaps.append(gap_headers)
    for cell in ws_gaps[1]:
        cell.font = bold
    gaps = summary.get("gaps", [])
    row_idx = 1
    for bucket in ("short", "mid", "long"):
        bucket_gaps = [g for g in gaps if g.get("feasibility") == bucket]
        if not bucket_gaps:
            continue
        ws_gaps.append([f"{FEASIBILITY_LABELS[bucket]} — {len(bucket_gaps)} item"
                        f"{'' if len(bucket_gaps) == 1 else 's'}"])
        row_idx += 1
        header_cell = ws_gaps.cell(row=row_idx, column=1)
        header_cell.font = bold
        for col in range(1, len(gap_headers) + 1):
            ws_gaps.cell(row=row_idx, column=col).fill = fill(_XLSX_FEAS_FILLS[bucket])
        for g in bucket_gaps:
            tier = g.get("tier")
            ws_gaps.append([_guard(v) for v in [
                g.get("rank"), g.get("technique_id"), g.get("name"),
                f"P{tier}" if isinstance(tier, int) and tier <= 3 else "Unranked",
                STATE_LABELS.get(g.get("state"), g.get("state")),
                (f"Uses logs you already collect: {g.get('via')}" if bucket == "short"
                 else f"Onboard from tooling you own: {g.get('via')}" if bucket == "mid" and g.get("via")
                 else g.get("via") or "Needs a new log source"),
                gap_recs.get(g.get("technique_id")) or g.get("hint"),
            ]])
            row_idx += 1
            if isinstance(tier, int) and tier in _XLSX_TIER_FILLS:
                ws_gaps.cell(row=row_idx, column=4).fill = fill(_XLSX_TIER_FILLS[tier])
            state_color = _XLSX_STATE_FILLS.get(g.get("state"))
            if state_color:
                ws_gaps.cell(row=row_idx, column=5).fill = fill(state_color)
    for row in ws_gaps.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    ws_gaps.freeze_panes = "A2"
    for i, width in enumerate([7, 12, 34, 12, 12, 34, 70], start=1):
        ws_gaps.column_dimensions[get_column_letter(i)].width = width

    # ------------------------------------------------------------------ Roadmap
    sheet(
        "Roadmap",
        ["Bucket", "Technique", "Name", "Action"],
        [
            [FEASIBILITY_LABELS[b], g.get("technique_id"), g.get("name"), g.get("hint")]
            for b in ("short", "mid", "long")
            for g in summary.get("roadmap", {}).get(b, [])
        ],
        [18, 12, 34, 70],
    )

    sheet(
        "Not Applicable",
        ["Technique", "Matrix", "Reason"],
        [
            [n.get("technique_id"), DOMAIN_LABELS.get(n.get("domain"), n.get("domain")), n.get("reason")]
            for n in summary.get("not_applicable", [])
        ],
        [12, 10, 80],
    )

    sheet(
        "Assumptions",
        ["Assumption"],
        [[a] for a in summary.get("assumptions", [])],
        [110],
    )

    # Phase 14g: per-entry environment evidence trail (present for
    # assessments created after the parser gained interpretations).
    interpretations = (params.get("environment_lists") or {}).get("interpretations") or []
    if interpretations:
        sheet(
            "How We Read Your Files",
            ["Your entry", "Sheet", "How it was read"],
            [
                [i.get("entry"), i.get("sheet"), i.get("interpretation")]
                for i in interpretations
            ],
            [40, 18, 80],
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
