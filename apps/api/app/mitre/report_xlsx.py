"""MITRE assessment XLSX gap-register builder (split out of report.py in
Phase 14h so the PDF/HTML builder and the spreadsheet builder are
independently editable; router.py still calls it via
``app.mitre.report.build_xlsx_export`` — report.py re-exports this module's
public names unchanged).

Computed numbers come ONLY from the stored summary/technique_results JSONB —
never recomputed here. Every customer/LLM string goes through `_guard` (Excel
formula-injection guard): rule names, descriptions, exclusion reasons, and
narrative output are all attacker-controlled.
"""

import io
import re

from app.mitre import attack_data
from app.mitre.report_common import (
    DOMAIN_LABELS,
    FEASIBILITY_LABELS,
    _MAPPING_STATUS_PLAIN_XLSX,
    _ordered_domains,
    _row_ref_sort_key,
    compute_log_source_coverage,
    resolve_branding,
)


def _guard(value):
    """Excel formula-injection guard (plan §10): a leading =, +, - or @ in an
    attacker-controlled string (rule names, descriptions, reasons) would
    execute as a formula when the register opens in Excel."""
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        return "'" + value
    return value


# ---- Assumptions tab: plain-language presentation (2026-08-18) ------------
# Deterministic (never LLM): each stored assumption string gets an Area and a
# "what this means for you" sentence in human words, matched by substring.
# Order matters twice — first match wins, and the sheet sorts by this order.
_ASSUMPTION_AREAS = [
    (("has been restructured",), "ATT&CK framework update",
     "MITRE reorganized this technique in the current ATT&CK version. Your "
     "detection still counts — it is simply shown under the new technique "
     "ID. This is a framework update, not a detection or security gap."),
    (("no longer maintains", "deprecated"), "ATT&CK framework update",
     "MITRE retired this technique ID, so it is noted for reference and "
     "left out of the score. A framework change, not a problem with the rule."),
    (("not a valid ATT&CK",), "Data quality",
     "This tag doesn't match any current ATT&CK technique, so it was left "
     "out. Correct the tag in the source rule if the mapping matters."),
    (("pulled read-only",), "Data source",
     "The rules came straight from your SIEM over a read-only connection, "
     "so the list reflects exactly what was live at pull time."),
    (("auto-imported",), "Data source",
     "Log sources were read from your SIEM's own connector inventory. "
     "Upload an environment workbook if anything is missing."),
    (("not recognized",), "Data source",
     "These SIEM connectors aren't in the mapping table yet, so they were "
     "not counted as log sources. They can be added to the environment "
     "workbook by hand."),
    (("AI-tagged", "AI-suggested", "confidence floor", "emitted by the AI",
      "AI-extracted", "processed by the AI", "sections were scanned"),
     "AI tagging",
     "AI helped map or extract these rules; every number stays "
     "deterministic. Worth spot-checking the flagged rows in the Use-Case "
     "Mappings sheet before relying on them."),
    (("disabled",), "Scoring policy",
     "How disabled rules count toward coverage follows the policy chosen "
     "at intake — re-run with the other policy to compare."),
    (("scope exclusion",), "Scope",
     "A scope exclusion you requested was applied (or could not be "
     "matched). Excluded items never count against your score."),
    (("inventory", "no recognizable", "platform filtering", "lower bound",
      "platforms"), "Scope",
     "Without a complete environment inventory the assessment cannot rule "
     "techniques out, so the score reads lower than reality — never higher."),
    (("Log Sources sheet",), "Data quality",
     "A rule points at telemetry your inventory doesn't declare — either "
     "the inventory sheet is incomplete or the rule may no longer be "
     "receiving data."),
    (("limit", "stopped at", "only the", "only its"), "Scope",
     "The input was larger than the processing cap, so items beyond the "
     "cap are not included in the numbers."),
]
_AREA_ORDER = {}
for _probes, _area, _meaning in _ASSUMPTION_AREAS:
    _AREA_ORDER.setdefault(_area, len(_AREA_ORDER))


def _classify_assumption(text: str) -> tuple:
    low = str(text).lower()
    for probes, area, meaning in _ASSUMPTION_AREAS:
        if any(p.lower() in low for p in probes):
            return area, meaning
    return "General", ""


def _successor_name(tid: str) -> str:
    info = attack_data.DEFAULT.get(tid) or {}
    return f" ({info['name']})" if info.get("name") else ""


def _rewrite_legacy_revoked(text: str) -> str:
    """Assessments stored before 2026-08-18 carry 'revoked … remapped'
    wording; rewrite at render time so old workbooks read customer-friendly
    too (feedback: 'revoked' sounds like an error, not a framework update)."""
    text = re.sub(
        r"mapping (T\d{4}(?:\.\d{3})?) on (.+?) is revoked in ATT&CK "
        r"v([\w.]+) — remapped to (T\d{4}(?:\.\d{3})?)",
        lambda m: (
            f"MITRE ATT&CK update: {m.group(1)} (tagged on {m.group(2)}) has "
            f"been restructured and is now represented under "
            f"{m.group(4)}{_successor_name(m.group(4))} in ATT&CK "
            f"v{m.group(3)} — the detection counts toward {m.group(4)}; this "
            "is a framework update, not a gap"
        ),
        text,
    )
    text = re.sub(
        r"tag '([^']+)' is revoked in ATT&CK v([\w.]+) — "
        r"remapped to (T\d{4}(?:\.\d{3})?)",
        lambda m: (
            f"MITRE ATT&CK update: tag '{m.group(1)}' has been restructured "
            f"and is now represented under "
            f"{m.group(3)}{_successor_name(m.group(3))} in ATT&CK v{m.group(2)}"
        ),
        text,
    )
    return text


# Phase 14c: fill colors (ARGB-less hex; light fills, dark text stays legible)
_XLSX_STATE_FILLS = {
    "covered": "C6EFCE",
    "partial": "FFE699",
    "not_covered": "FFC7CE",
    "not_applicable": "D9D9D9",
}
_XLSX_FEAS_FILLS = {"short": "C6EFCE", "mid": "FFE699", "long": "D9D9D9"}
_STATE_PLAIN_XLSX = {
    "covered": "A rule detects this",
    "partial": "Half-covered",
    "not_covered": "No rule detects this",
    "not_applicable": "Doesn't apply to this environment",
}

# ---------------------------------------------------------------------------
# Reference KQL (2026-08-14, from the VFQ customer-deliverable review): an
# illustrative per-table skeleton for each buildable gap, so detection
# engineers see HOW to approach the query — with the two failure modes called
# out up front (false positives from missing allowlists/thresholds; rules
# that never fire because the table is empty). Deterministic; clearly
# labelled reference-only; NOT a production rule.
# ---------------------------------------------------------------------------
_KQL_HEADER = (
    "// REFERENCE ONLY - illustrative skeleton, not a production rule.\n"
    "// 1) Confirm data first:  {table} | take 10   (a rule on an empty\n"
    "//    table never fires). 2) Run in audit mode 1-2 weeks, baseline\n"
    "//    normal volume. 3) Add allowlists for known-good accounts/hosts\n"
    "//    BEFORE enabling alerts - thresholds beat single-event alerts.\n"
)

_KQL_TEMPLATES = {
    "SecurityEvent": "SecurityEvent\n| where TimeGenerated > ago(1h)\n| where // <condition for {tid} {name} - e.g. EventID + fields>\n| summarize Count=count() by Computer, Account, bin(TimeGenerated, 15m)\n| where Count > 5 // tune from your baseline",
    "Syslog": "Syslog\n| where TimeGenerated > ago(1h)\n| where Facility in ('auth','authpriv') // scope narrow first\n| where SyslogMessage has_any (/* markers for {tid} {name} */)\n| summarize Count=count() by HostName, bin(TimeGenerated, 15m)\n| where Count > 5",
    "DeviceProcessEvents": "DeviceProcessEvents\n| where TimeGenerated > ago(1h)\n| where // <process/command-line condition for {tid} {name}>\n| project TimeGenerated, DeviceName, AccountName, FileName,\n          ProcessCommandLine, InitiatingProcessFileName\n// FP control: exclude software-deployment and admin-tool accounts",
    "DeviceFileEvents": "DeviceFileEvents\n| where TimeGenerated > ago(1h)\n| where ActionType == 'FileCreated'\n| where // <path/extension/hash condition for {tid} {name}>\n| summarize Files=count() by DeviceName, InitiatingProcessAccountName, bin(TimeGenerated, 10m)\n| where Files > 10 // burst behaviour, not single files",
    "DeviceNetworkEvents": "DeviceNetworkEvents\n| where TimeGenerated > ago(1h)\n| where RemoteIPType == 'Public'\n| where // <port/protocol/destination condition for {tid} {name}>\n| summarize Conns=count() by DeviceName, RemoteIP, RemotePort, bin(TimeGenerated, 15m)\n| where Conns > 20 // beaconing/burst threshold - baseline first",
    "DeviceLogonEvents": "DeviceLogonEvents\n| where TimeGenerated > ago(1h)\n| where ActionType == 'LogonFailed'\n| summarize Failures=count(), Targets=dcount(DeviceName) by AccountName, bin(TimeGenerated, 15m)\n| where Failures > 20 or Targets > 5 // spray pattern for {tid} {name}",
    "SigninLogs": "SigninLogs\n| where TimeGenerated > ago(1h)\n| where ResultType == 0 // successes - failures are a separate rule\n| where // <app / location / device condition for {tid} {name}>\n| summarize by UserPrincipalName, IPAddress, Location, AppDisplayName\n// FP control: suppress corporate egress IPs and travel-approved users",
    "AuditLogs": "AuditLogs\n| where TimeGenerated > ago(1h)\n| where OperationName has_any (/* directory operations for {tid} {name} */)\n| project TimeGenerated, OperationName, InitiatedBy, TargetResources\n// FP control: exclude your IAM automation service principals",
    "AzureActivity": "AzureActivity\n| where TimeGenerated > ago(1h)\n| where OperationNameValue has_any (/* operations for {tid} {name} */)\n| where ActivityStatusValue == 'Success'\n| project TimeGenerated, Caller, OperationNameValue, ResourceGroup\n// FP control: exclude IaC pipeline identities (alert if they act off-schedule)",
    "AzureDiagnostics": "AzureDiagnostics\n| where TimeGenerated > ago(1h)\n| where Category == '<pick the one category you need>' // never query the whole table\n| where // <condition for {tid} {name}>\n| summarize Count=count() by Resource, bin(TimeGenerated, 15m)",
    "CommonSecurityLog": "CommonSecurityLog\n| where TimeGenerated > ago(1h)\n| where // <vendor/activity/port/direction condition for {tid} {name}>\n| summarize Count=count() by SourceIP, DestinationIP, DestinationPort, bin(TimeGenerated, 15m)\n| where Count > 10",
    "OfficeActivity": "OfficeActivity\n| where TimeGenerated > ago(1h)\n| where Operation has_any (/* operations for {tid} {name} */)\n| summarize Events=count() by UserId, Operation, bin(TimeGenerated, 15m)\n| where Events > 20 // bulk behaviour, not single clicks",
    "EmailAttachmentInfo": "EmailAttachmentInfo\n| where TimeGenerated > ago(1h)\n| join kind=inner (EmailEvents | where TimeGenerated > ago(1h)) on NetworkMessageId\n| where // <file-type / sender condition for {tid} {name}>\n| project TimeGenerated, SenderFromAddress, RecipientEmailAddress, FileName, FileType, SHA256",
}
_KQL_GENERIC = (
    "{table}\n| where TimeGenerated > ago(1h)\n| where // <condition for {tid} {name}>\n"
    "| summarize Count=count() by bin(TimeGenerated, 15m)\n| where Count > 5 // tune from baseline"
)


def _reference_kql(tid, name, via):
    """Illustrative KQL skeleton for a buildable gap, or None when the gap
    has no named source (long-term / bespoke gaps)."""
    table = str(via or "").replace("Sentinel table - ", "").strip()
    if not table or " " in table:
        # no via, or a prose/multi-word source name that isn't a KQL table
        return None
    template = _KQL_TEMPLATES.get(table, _KQL_GENERIC)
    return _KQL_HEADER.format(table=table) + template.format(
        table=table, tid=tid, name=str(name or "")[:40]
    )


def build_xlsx_export(assessment, use_cases: list, scope: str = "full",
                       branding: dict | None = None) -> bytes:
    """The detailed gap register as a workbook (Phase 14c polish, Phase A9
    consolidation): 'Read Me' guide sheet first, colored state/priority/
    feasibility cells, frozen headers + auto-filter + wrapped text
    everywhere, technique names + plain-words 'Why' column, numerically
    sorted rule rows. Computed numbers are untouched — styling and wording
    only. Phase A9 merged the old 'Technique Register', 'Gaps &
    Recommendations', and 'Roadmap' sheets (Roadmap was the same gap dicts
    re-bucketed; Gaps was a subset of the Register — pure duplication) into
    ONE 'Technique Tracker' sheet: one row per applicable technique, gap-only
    columns blank for covered rows, plus blank Owner/Status/Target date/Notes
    columns so it doubles as a working tracker. 'Log fields needed' is
    curated per data-source component via plain_language.telemetry_lines;
    blank for techniques with no curated component. Phase A10 piece 3 adds
    'Coverage by Log Source' — one row per detection-rule log source
    (report_common.compute_log_source_coverage, shared with the results-
    page drill-down so the two views can't drift), excluded from scoped
    downloads the same way 'Use-Case Mappings' already is.

    branding: optional org overrides (display name/accent/watermark — plan
    §14h). Resolved here for a consistent shape across callers; workbook
    core properties (title/author/company) start consuming it in the
    XLSX-polish follow-up.
    """
    branding = resolve_branding(branding)
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    from app.mitre import attack_data, plain_language, ranking

    summary = assessment.summary or {}
    params = assessment.params or {}
    overall = summary.get("overall", {})
    narrative = summary.get("narrative", {})
    gap_recs = narrative.get("gap_recommendations", {})
    index = attack_data.DEFAULT

    wb = Workbook()
    bold = Font(bold=True)
    italic = Font(italic=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    wrap_center = Alignment(wrap_text=True, vertical="top", horizontal="center")
    # Clearly visible grid (the earlier pale blue read as "no border" in Excel)
    thin = Side(style="thin", color="B9AECB")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    # Phase A11: ONE consistent branded header fill/font, audited and made
    # uniform across every sheet. 2026-08-14 (VFQ customer-deliverable
    # restyle): deep-purple headers + teal section bands + zebra data rows —
    # the same palette the PPTX briefing deck (report_pptx.py) uses.
    BRAND = "341954"
    ACCENT = "00A98B"
    ZEBRA = "F3F0F7"
    white_bold = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14, color="FFFFFF")

    def fill(color):
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def sheet(title, headers, rows, widths, first=False, filters=True,
              borders=True, center_cols=()):
        ws = wb.active if first else wb.create_sheet()
        ws.title = title
        ws.append([_guard(h) for h in headers])
        for cell in ws[1]:
            cell.font = white_bold
            cell.fill = fill(BRAND)
            if borders:
                cell.border = cell_border
            if cell.column in center_cols:
                cell.alignment = wrap_center
        for row in rows:
            ws.append([_guard(v) for v in row])
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap_center if cell.column in center_cols else wrap
                if borders:
                    cell.border = cell_border
                if cell.row % 2 == 0:  # zebra rows (state/tier fills overwrite)
                    cell.fill = fill(ZEBRA)
        ws.freeze_panes = "A2"
        if filters and rows:
            ws.auto_filter.ref = ws.dimensions
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        return ws

    # ---------------------------------------------------------------- Read Me
    strict_pct = overall.get("strict_pct")
    readme_rows = [
        ["This workbook is the full detail behind your MITRE ATT&CK coverage assessment."],
        [],
        ["The three key numbers"],
        [f"Coverage: {strict_pct}%",
         f"{overall.get('covered')} of {overall.get('applicable')} techniques that "
         "apply to your environment have at least one detection rule."],
        [f"Gaps to work on: {len(summary.get('gaps', []))}",
         "Ranked by priority in the 'Technique Tracker' sheet — start at the top."],
        [f"Rules analyzed: {(summary.get('counts') or {}).get('use_cases', len(use_cases))}",
         "Your uploaded detection rules — see 'Use-Case Mappings' for what each one maps to."],
        [f"Is {strict_pct}% bad? Probably not: early SIEM detection programs typically "
         "start under 10%, because ATT&CK counts every known attacker technique. "
         "The roadmap matters more than the grade."],
        [],
        ["What each sheet contains"],
        ["Summary", "The headline numbers with a plain-words explanation of each."],
        ["Coverage by Tactic", "Coverage per attack stage (tactic), per matrix."],
        ["Technique Tracker", "One row per applicable technique: state, why, priority, and — "
         "for gaps — the recommendation, roadmap bucket, and blank Owner/Status/Target date/"
         "Notes columns for you to fill in as you work the plan."],
        ["Use-Case Mappings", "Your rules, one per row, with how each was mapped."],
        ["Coverage by Log Source", "What each log source (e.g. Sysmon, CloudTrail) buys you: "
         "how many rules use it and which techniques/attack stages those rules cover."],
        ["Not Applicable", "Techniques that don't count toward your score, with reasons."],
        ["Assumptions", "Context behind the numbers, in plain words — what shaped "
         "them, what changed, and what (if anything) to do about it."],
        [],
        ["Color legend"],
        ["Covered", "A rule detects this technique."],
        ["Partial", "Only a disabled rule, a low-confidence mapping, or a sub-technique reaches it."],
        ["Not covered", "No rule detects it — this is a gap."],
        ["N/A", "Doesn't apply to your environment; excluded from the score."],
    ]
    ws_readme = sheet("Read Me", ["How to read this workbook", ""],
                      readme_rows, [46, 100], first=True, filters=False,
                      borders=False)  # prose sheet — a full grid would be noise
    for label, color in (("Covered", "covered"), ("Partial", "partial"),
                         ("Not covered", "not_covered"), ("N/A", "not_applicable")):
        for row in ws_readme.iter_rows(min_row=2, max_col=1):
            if row[0].value == label:
                row[0].fill = fill(_XLSX_STATE_FILLS[color])
    for row in ws_readme.iter_rows(min_row=2, max_col=1):
        if row[0].value in ("The three key numbers", "What each sheet contains", "Color legend"):
            row[0].font = bold
    ws_readme.protection.sheet = True  # guide sheet is read-only; no password (accidental-edit guard only)

    # ------------------------- Summary (redesigned post-14: sectioned, ----
    # ------------------------- colored, with an executive summary) --------

    def _pct_fill_color(pct):
        value = float(pct or 0)
        if value >= 50:
            return _XLSX_STATE_FILLS["covered"]
        if value >= 15:
            return _XLSX_STATE_FILLS["partial"]
        return _XLSX_STATE_FILLS["not_covered"]

    ws_sum = wb.create_sheet()
    ws_sum.title = "Summary"
    for i, w in enumerate((28, 46, 78), start=1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w

    def sum_row(values, *, fills=None, fonts=None, merge=False, height=None,
                center=()):
        ws_sum.append([_guard(v) for v in values])
        r = ws_sum.max_row
        if merge:
            ws_sum.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        for c in range(1, 4):
            cell = ws_sum.cell(row=r, column=c)
            cell.alignment = wrap_center if c in center else wrap
            cell.border = cell_border  # all-borders, merged content included
            if fills and fills.get(c):
                cell.fill = fill(fills[c])
            if fonts and fonts.get(c):
                cell.font = fonts[c]
        if height:
            ws_sum.row_dimensions[r].height = height
        return r

    def section(title):
        sum_row([""])
        sum_row([title, "", ""], merge=True,
                fills={1: ACCENT, 2: ACCENT, 3: ACCENT}, fonts={1: white_bold})

    intake = params.get("intake") or {}
    sum_row(["MITRE ATT&CK Coverage Assessment", "", ""], merge=True,
            fills={1: BRAND, 2: BRAND, 3: BRAND}, fonts={1: title_font}, height=26)
    subtitle = assessment.name
    if intake.get("project_name"):
        subtitle = f"{intake['project_name']} — {assessment.name}"
    sum_row([subtitle, "", ""], merge=True, fonts={1: bold})
    sum_row([f"ATT&CK v{assessment.attack_version} · run completed "
             f"{str(assessment.completed_at or '')[:16]}", "", ""], merge=True)

    # Simple compact pointers instead of a paragraph — deterministic, built
    # from the computed numbers (the fuller narrative stays in the PDF).
    section("EXECUTIVE SUMMARY")
    top_gap_names = ", ".join(
        f"{g.get('technique_id')} {g.get('name')}" for g in summary.get("gaps", [])[:3]
    )
    roadmap_counts = {b: len(summary.get("roadmap", {}).get(b, [])) for b in ("short", "mid", "long")}
    pointers = [
        (f"• Your rules detect {overall.get('covered')} of "
         f"{overall.get('applicable')} applicable attacker techniques — "
         f"{overall.get('strict_pct')}% coverage ({overall.get('weighted_pct')}% weighted).",
         bold),
        (f"• {overall.get('not_covered')} techniques have no detection today; "
         f"{overall.get('partial')} are only partially covered.", None),
    ]
    if top_gap_names:
        pointers.append((f"• Start here: {top_gap_names}.", bold))
    pointers.append(
        (f"• Roadmap: {roadmap_counts['short']} gaps buildable now with logs you "
         f"already collect · {roadmap_counts['mid']} need log onboarding first · "
         f"{roadmap_counts['long']} need a new capability.", None))
    # 2026-08-13 VFQ review: "buildable now" only says the source is DECLARED
    # onboarded — when many existing rules carry Last Triggered "never", the
    # same sources demonstrably haven't produced detections yet.
    never_count = sum(
        1 for uc in use_cases
        if str(uc.get("last_triggered") or "").strip().lower() == "never"
    )
    if never_count and roadmap_counts["short"]:
        pointers.append(
            (f"• Caveat: 'buildable now' means the log source is declared as "
             f"onboarded — {never_count} of {len(use_cases)} rules found no "
             "events when last validated (Last Triggered: never), so verify "
             "source health before committing build dates.", italic))
    pointers.append(
        (f"• Is {overall.get('strict_pct')}% bad? Probably not: early SIEM "
         "programs typically start under 10% — the roadmap matters more than "
         "the grade.", italic))
    for text, font in pointers:
        sum_row([text, "", ""], merge=True, height=26,
                fonts={1: font} if font else None)

    section("KEY NUMBERS")
    sum_row(["Metric", "Value", "What it means"],
            fills={1: BRAND, 2: BRAND, 3: BRAND},
            fonts={1: white_bold, 2: white_bold, 3: white_bold}, center=(2,))
    r = sum_row(["Coverage %", overall.get("strict_pct"),
                 "Of the techniques that apply to your environment, the share "
                 "with at least one qualifying detection rule."], center=(2,))
    ws_sum.cell(row=r, column=2).fill = fill(_pct_fill_color(overall.get("strict_pct")))
    ws_sum.cell(row=r, column=1).font = bold
    ws_sum.cell(row=r, column=2).font = bold
    r = sum_row(["Weighted coverage %", overall.get("weighted_pct"),
                 "Same, but half-covered techniques count as 0.5 instead of 0."],
                center=(2,))
    ws_sum.cell(row=r, column=2).font = bold
    for label, key, meaning in (
        ("Covered", "covered", "Techniques at least one enabled rule detects."),
        ("Partial", "partial", "Techniques reached only by a disabled rule, a "
         "low-confidence mapping, or a covered sub-technique."),
        ("Not covered", "not_covered", "Techniques no rule detects — the gaps."),
        ("Not applicable", "not_applicable", "Techniques excluded from the score "
         "(wrong platform, excluded by you, or deprecated)."),
    ):
        r = sum_row([label, overall.get(key), meaning], center=(2,))
        ws_sum.cell(row=r, column=2).fill = fill(_XLSX_STATE_FILLS[key])
        ws_sum.cell(row=r, column=2).font = bold
    sum_row(["Applicable techniques", overall.get("applicable"),
             "The denominator: techniques that apply to your environment."],
            center=(2,))
    for k, d in _ordered_domains(summary.get("domains")):
        r = sum_row([f"{DOMAIN_LABELS.get(k, k)} coverage %", d.get("strict_pct"),
                     f"Coverage within the {DOMAIN_LABELS.get(k, k)} matrix only."],
                    center=(2,))
        if d.get("applicable"):
            ws_sum.cell(row=r, column=2).fill = fill(_pct_fill_color(d.get("strict_pct")))

    top_gaps = summary.get("gaps", [])[:5]
    if top_gaps:
        section("TOP 5 THINGS TO FIX FIRST")
        sum_row(["Gap", "Effort", "Recommendation"],
                fills={1: BRAND, 2: BRAND, 3: BRAND},
                fonts={1: white_bold, 2: white_bold, 3: white_bold}, center=(2,))
        for g in top_gaps:
            r = sum_row([
                f"#{g.get('rank')} {g.get('technique_id')} {g.get('name')}",
                FEASIBILITY_LABELS.get(g.get("feasibility"), ""),
                gap_recs.get(g.get("technique_id")) or g.get("hint") or "",
            ], fonts={1: bold}, center=(2,))
            bucket_color = _XLSX_FEAS_FILLS.get(g.get("feasibility"))
            if bucket_color:
                ws_sum.cell(row=r, column=2).fill = fill(bucket_color)

    section("ABOUT THIS ASSESSMENT")
    sum_row(["Assessment", assessment.name, "The name this run was saved under."])
    for key, label, meaning in (
        ("project_name", "Organization / project", "Who this assessment is for."),
        ("scope_label", "Scope", "The department or scope this run covers."),
        ("prepared_by", "Prepared by", "Who prepared this assessment."),
        ("purpose_note", "Purpose", "Why this assessment was run."),
    ):
        if intake.get(key):
            sum_row([label, intake[key], meaning])
    sum_row(["ATT&CK version", assessment.attack_version,
             "The MITRE ATT&CK release the assessment is pinned to."])
    sum_row(["Narrative", narrative.get("generated_by", ""),
             "Whether recommendation wording was AI-written or standard template "
             "text (numbers are computed either way)."])
    sum_row(["Thresholds", str(params.get("thresholds", {})),
             "The confidence cut-offs used to count a mapping as coverage."])
    ws_sum.freeze_panes = "A2"

    # ------------------------------------------------------ Coverage by Tactic
    tactic_rows = [
        [DOMAIN_LABELS.get(dk, dk), t.get("name"), t.get("covered"), t.get("partial"),
         t.get("not_covered"), t.get("not_applicable"), t.get("applicable"),
         t.get("strict_pct"), t.get("weighted_pct")]
        for dk, d in _ordered_domains(summary.get("domains"))
        for t in d.get("tactics", [])
    ]
    ws_tactic = sheet(
        "Coverage by Tactic",
        ["Matrix", "Tactic", "Covered", "Partial", "Not covered", "N/A", "Applicable", "Coverage %", "Weighted %"],
        tactic_rows,
        [12, 26, 10, 10, 12, 8, 12, 12, 12],
        center_cols=(3, 4, 5, 6, 7, 8, 9),
    )
    if tactic_rows:
        last_tactic_row = 1 + len(tactic_rows)
        ws_tactic.conditional_formatting.add(
            f"H2:H{last_tactic_row}",
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color="638EC6"),
        )
        ws_tactic.conditional_formatting.add(
            f"I2:I{last_tactic_row}",
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color="A9C4EB"),
        )
        chart = BarChart()
        chart.type = "col"
        chart.title = "Coverage % by tactic"
        chart.y_axis.title = "Coverage %"
        chart.x_axis.title = "Tactic"
        chart.height, chart.width = 10, 24
        data = Reference(ws_tactic, min_col=8, min_row=1, max_row=last_tactic_row)
        cats = Reference(ws_tactic, min_col=2, min_row=2, max_row=last_tactic_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws_tactic.add_chart(chart, "K2")

    # --------------------------------------------------------- Technique Tracker
    mapped_by_technique: dict = {}
    for uc in use_cases:
        for m in uc.get("mappings") or []:
            tid = m.get("technique_id")
            mapped_by_technique.setdefault(tid, []).append(
                {"name": uc.get("name"), "enabled": uc.get("enabled"),
                 "source": m.get("source"), "confidence": m.get("confidence")}
            )
    results = assessment.technique_results or []
    state_by_id = {r.get("technique_id"): r.get("state") for r in results}
    total_rules = (summary.get("counts") or {}).get("use_cases", len(use_cases))
    tactic_names = {
        (domain, t["id"]): t["name"]
        for domain in index.domains
        for t in index.tactics(domain)
    }
    confidence_covered = (params.get("thresholds") or {}).get("confidence_covered", 0.7)
    gaps_by_id = {g.get("technique_id"): g for g in summary.get("gaps", [])}
    _ROADMAP_BUCKET_VALUE = {"short": "Short", "mid": "Mid", "long": "Long"}
    # Applicable = every state EXCEPT not_applicable (matches coverage.py's
    # own "applicable" denominator) — N/A techniques stay in their own sheet.
    # 2026-08-14 (VFQ review): action-first ordering — gaps at the top in
    # their ranked order, covered rows after, so the sheet reads as a work
    # queue instead of a register (auto-filter still allows any re-sort).
    _TRACKER_STATE_ORDER = {"not_covered": 0, "partial": 1, "covered": 2}
    applicable_results = sorted(
        (r for r in results if r.get("state") != "not_applicable"),
        key=lambda r: (
            _TRACKER_STATE_ORDER.get(r.get("state"), 3),
            (gaps_by_id.get(r.get("technique_id")) or {}).get("rank") or 10 ** 6,
            str(r.get("technique_id")),
        ),
    )

    def tracker_row(r):
        tid = r.get("technique_id")
        mapped = mapped_by_technique.get(tid, [])
        why = plain_language.derive_why(
            r, mapped,
            total_rules=total_rules,
            sub_states=plain_language.sub_states_for(r, mapped, state_by_id, index),
            confidence_covered=confidence_covered,
        )
        gap = gaps_by_id.get(tid)  # None for covered rows — gap-only columns stay blank
        tier = gap.get("tier") if gap else None
        ranked = isinstance(tier, int) and tier <= 3
        relevance = (gap or {}).get("threat_relevance") or []
        telemetry_cell = ""
        if gap:
            lines = plain_language.telemetry_lines(tid, index)
            # 2026-08-14 (VFQ review): lead with the component that matches
            # the gap's chosen telemetry category, so the field guidance
            # agrees with the recommended source instead of raw ATT&CK order.
            category = gap.get("category")
            if category:
                lines = sorted(
                    lines,
                    key=lambda ln: 0 if ranking.component_category(
                        ln.split(" — ")[0]) == category else 1,
                )
            telemetry_cell = "\n\n".join(lines)
        recommendation = (gap_recs.get(tid) or gap.get("hint") or "") if gap else ""
        return [
            tid,
            (index.get(tid) or {}).get("name", ""),
            ", ".join(tactic_names.get((r.get("domain"), t), t) for t in r.get("tactics", [])),
            DOMAIN_LABELS.get(r.get("domain"), r.get("domain")),
            _STATE_PLAIN_XLSX.get(r.get("state"), ""),
            why,
            r.get("strength"),
            (tier if ranked else "Unranked") if gap else None,
            ", ".join(relevance) if relevance else "",
            "Yes" if (gap or {}).get("crown_jewel_relevant") else "",
            FEASIBILITY_LABELS.get((gap or {}).get("feasibility"), "") if gap else "",
            _ROADMAP_BUCKET_VALUE.get((gap or {}).get("feasibility"), "") if gap else "",
            recommendation,
            telemetry_cell,
            _reference_kql(tid, (index.get(tid) or {}).get("name"),
                           (gap or {}).get("via")) if gap else "",
            (gap or {}).get("via") or "" if gap else "",
            "", "", "", "",  # Owner, Status, Target date, Notes — blank tracking columns
        ]

    ws_tracker = sheet(
        "Technique Tracker",
        ["Technique ID", "Name", "Tactic(s)", "Domain", "State", "Why", "Strength",
         "Priority", "Threat match", "Crown jewel", "Feasibility", "Roadmap bucket",
         "Recommendation", "Log fields needed",
         "Reference KQL (illustrative — tune before use)", "Via",
         "Owner", "Status", "Target date", "Notes"],
        [tracker_row(r) for r in applicable_results],
        [12, 30, 22, 10, 22, 55, 10, 10, 22, 10, 22, 14, 55, 45, 55, 22, 16, 14, 14, 30],
        center_cols=(4, 7, 8, 10, 12),
    )
    for i, r in enumerate(applicable_results, start=2):
        color = _XLSX_STATE_FILLS.get(r.get("state"))
        if color:
            ws_tracker.cell(row=i, column=5).fill = fill(color)
        priority_cell = ws_tracker.cell(row=i, column=8)
        if isinstance(priority_cell.value, int):
            # numeric so the color scale below can rank it; displays as "P1"/"P2"/"P3"
            priority_cell.number_format = '"P"0'
    if applicable_results:
        last_tracker_row = 1 + len(applicable_results)
        ws_tracker.conditional_formatting.add(
            f"H2:H{last_tracker_row}",
            ColorScaleRule(start_type="min", start_color="F8696B",
                          mid_type="percentile", mid_value=50, mid_color="FFEB84",
                          end_type="max", end_color="63BE7B"),
        )

    # -------------------------------------------------------- Use-Case Mappings
    sorted_ucs = sorted(use_cases, key=_row_ref_sort_key)
    sheet(
        "Use-Case Mappings",
        ["Row", "Rule name", "Enabled", "How it was mapped", "Techniques", "Confidence", "Source", "Log source", "Description", "Logic"],
        [
            [uc.get("row_ref"), uc.get("name"),
             {True: "yes", False: "no"}.get(uc.get("enabled"), "unknown"),
             _MAPPING_STATUS_PLAIN_XLSX.get(uc.get("mapping_status"), uc.get("mapping_status")),
             ", ".join(m.get("technique_id", "") for m in (uc.get("mappings") or [])),
             ", ".join(str(m.get("confidence", "")) for m in (uc.get("mappings") or [])),
             ", ".join(m.get("source", "") for m in (uc.get("mappings") or [])),
             uc.get("log_source") or "", uc.get("description") or "",
             uc.get("logic") or ""]  # Phase 7; query text is attacker-controlled — _guard applies
            for uc in sorted_ucs
        ],
        [10, 42, 9, 30, 22, 12, 12, 16, 60, 60],
        center_cols=(3, 6),
    )

    # ------------------------------------------------------ Coverage by Log Source
    # Phase A10 piece 3: "see what each device buys you" — deterministic
    # read-time grouping, same function the results-page drill-down uses
    # (report_common.compute_log_source_coverage), so the sheet and the UI
    # can never drift apart.
    log_source_groups = compute_log_source_coverage(
        use_cases, assessment.technique_results or [], index
    )
    sheet(
        "Coverage by Log Source",
        ["Log source", "Rules", "Techniques covered", "Attack stages", "Techniques"],
        [
            [g["log_source"], g["rule_count"], g["techniques_covered"],
             ", ".join(g["tactics"]),
             ", ".join(f"{t['technique_id']} {t['name']}" for t in g["techniques"])]
            for g in log_source_groups
        ],
        [28, 10, 18, 40, 70],
        center_cols=(2, 3),
    )

    sheet(
        "Not Applicable",
        ["Technique", "Matrix", "Reason"],
        [
            [n.get("technique_id"), DOMAIN_LABELS.get(n.get("domain"), n.get("domain")), n.get("reason")]
            for n in summary.get("not_applicable", [])
        ],
        [12, 10, 80],
        center_cols=(2,),
    )

    assumption_rows = []
    for a in summary.get("assumptions", []):
        friendly = _rewrite_legacy_revoked(str(a))
        area, meaning = _classify_assumption(friendly)
        assumption_rows.append([area, friendly, meaning])
    # stable sort groups related notes together, in the classifier's order
    assumption_rows.sort(key=lambda r: _AREA_ORDER.get(r[0], len(_AREA_ORDER)))
    sheet(
        "Assumptions",
        ["Area", "Assumption", "What this means for you"],
        assumption_rows,
        [24, 78, 62],
    )

    # Phase 14g: per-entry environment evidence trail (present for
    # assessments created after the parser gained interpretations).
    interpretations = (params.get("environment_lists") or {}).get("interpretations") or []
    if interpretations:
        sheet(
            "How We Read Your Files",
            ["Your entry", "Sheet", "How it was read"],
            [
                [i.get("entry"), i.get("sheet"), i.get("interpretation")]
                for i in interpretations
            ],
            [40, 18, 80],
        )

    # Per-tab download: keep only that tab's sheets (built once, then pruned
    # — cheapest way to guarantee identical content and styling). The
    # Tracker sheet carries both coverage state and gap detail (Phase A9
    # merge), so both scopes keep it.
    _SCOPE_SHEETS = {
        "coverage": {"Coverage by Tactic", "Technique Tracker"},
        "gaps": {"Technique Tracker"},
        "assumptions": {"Not Applicable", "Assumptions", "How We Read Your Files"},
    }
    if scope in _SCOPE_SHEETS:
        for name in list(wb.sheetnames):
            if name not in _SCOPE_SHEETS[scope]:
                del wb[name]

    # openpyxl has no wired-up "Company" extended property (docProps/app.xml) —
    # only docProps/core.xml fields are settable, so the org name goes in
    # description instead.
    wb.properties.title = f"MITRE ATT&CK Coverage Assessment — {assessment.name}"
    wb.properties.creator = branding["report_display_name"] or "MITRE ATT&CK Coverage Assessment"
    if branding["report_display_name"]:
        wb.properties.description = f"Prepared for {branding['report_display_name']}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
