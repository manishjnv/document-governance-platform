"""Structured file ingest for the MITRE assessment module (Phase 1).

Reads use-case dumps (xlsx/xls/csv) and the multi-sheet environment workbook
with DIRECT cell access (openpyxl/xlrd/stdlib csv) — deliberately NOT
app.parser.parse_document(), whose ExcelParser flattens sheets to text and
destroys the column alignment TTP tags live in.

This is a trust boundary: callers hand it raw uploaded bytes. All guards
(size, row caps, empty parse, undetectable columns) raise IngestError; the
router maps IngestError -> HTTP 422.

pdf/docx dumps are rejected here in Phase 1 — AI text-extraction lands in
Phase 2 with the tagging agent.
"""

import csv
import io
import re

MAX_FILE_SIZE = 50 * 1024 * 1024  # matches documents.py
MAX_USE_CASE_ROWS = 5_000
MAX_ASSET_ROWS = 10_000
_HEADER_SCAN_ROWS = 10  # header row must appear in the first N rows
# Workbook-wide budgets (2026-08-01 adversarial review, blocking finding #2:
# the old per-sheet row cap let a many-sheet zip-bomb xlsx buffer unbounded
# memory in the synchronous create handler).
MAX_SHEETS = 20
MAX_TOTAL_ROWS = MAX_USE_CASE_ROWS + MAX_ASSET_ROWS + 5 * _HEADER_SCAN_ROWS  # cumulative, whole workbook
MAX_ROW_CELLS = 64

TAG_RE = re.compile(r"T\d{4}(?:\.\d{3})?", re.IGNORECASE)

PDF_DOCX_MESSAGE = (
    "PDF/DOCX rule dumps are enabled with AI extraction in an upcoming "
    "release — please use the XLSX template for now"
)
NO_NAME_COLUMN_MESSAGE = (
    "Could not detect a use-case name column in the uploaded file. Please "
    "use the ScopeWise use-case template (columns: use-case name, MITRE "
    "technique tags, detection logic, description, log source, status)."
)


class IngestError(ValueError):
    """Parse/validation failure the router reports as HTTP 422."""

    def __init__(self, detail: str):
        super().__init__(detail)
        self.detail = detail


def _norm(value) -> str:
    """Normalize a header/sheet-name cell for synonym matching."""
    text = str(value or "").lower().replace("_", " ")
    text = text.replace("(s)", "s")  # "Technique(s)" -> "techniques"
    text = re.sub(r"[^a-z0-9 ]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


# Field detection order matters: 'tags' before 'name' so "mitre technique"
# never lands on a name-ish synonym, 'logic' before 'description'.
# Entries are POST-_norm forms: "ATT&CK ID" -> "att ck id",
# "MITRE_TTP" -> "mitre ttp" (Phase 6 widened these over real SIEM exports:
# Splunk ES, Sentinel, Elastic, QRadar header variants).
COLUMN_SYNONYMS = {
    "tags": {
        "technique", "techniques", "technique id", "technique ids", "ttp",
        "ttps", "mitre", "mitre id", "mitre ids", "mitre technique",
        "mitre techniques", "mitre technique id", "mitre technique ids",
        "mitre technique tags", "mitre tags", "attack id", "attack ids",
        "attack technique", "attack techniques", "mitre attack",
        "mitre attack technique", "mitre attack techniques", "mitre attack id",
        "att ck", "att ck id", "att ck ids", "att ck technique",
        "att ck techniques", "mitre att ck", "mitre att ck id",
        "mitre att ck technique", "mitre att ck techniques", "mitre ttp",
        "mitre ttps", "ttp id", "ttp ids", "technique tag", "technique tags",
        "attack technique id", "attack technique ids", "mapped technique",
        "mapped techniques", "mitre mapping", "mitre mappings",
    },
    "name": {
        "name", "use case", "use cases", "use case name", "usecase",
        "usecase name", "rule", "rule name", "title", "detection",
        "detection name", "detection rule", "alert name", "signature name",
        "detection title", "rule title", "alert", "alert rule",
        "search name", "analytic", "analytic name", "content name",
        "detection rule name", "signature", "display name", "use case title",
        "correlation search", "correlation search name",
    },
    "logic": {
        "logic", "detection logic", "rule logic", "query", "search",
        "search query", "condition", "detection condition", "spl", "kql",
        "correlation rule", "detection condition logic",
        "detection query", "rule query", "query string", "search string",
        "aql", "eql", "esql", "sigma", "sigma rule", "rule definition",
        "detection rule logic", "alert query", "kql query", "spl query",
        "detection criteria", "criteria", "rule expression", "expression",
    },
    "description": {
        "description", "desc", "summary", "details", "notes",
        "rule description", "detection description", "objective", "purpose",
        "use case description", "comment", "comments",
    },
    "enabled": {
        "status", "enabled", "state", "active", "rule status", "enabled disabled",
        "is enabled", "enabled status", "rule state", "deployment status",
        "deployed", "production status", "active status", "enable",
    },
    "log_source": {
        "log source", "log sources", "logsource", "data source",
        "data sources", "datasource", "index", "source", "sourcetype",
        "log source index",
        "log type", "data type", "event source", "source type",
        "telemetry", "telemetry source", "source index", "device type",
        "log source type", "siem index",
    },
}

_TRUE_WORDS = {"enabled", "true", "yes", "active", "on", "1", "prod", "production", "live"}
_FALSE_WORDS = {"disabled", "false", "no", "inactive", "off", "0", "draft", "paused"}


def parse_enabled(value) -> bool | None:
    """Map a status cell to True/False, or None when unrecognizable."""
    if isinstance(value, bool):
        return value
    word = _norm(value)
    if word in _TRUE_WORDS:
        return True
    if word in _FALSE_WORDS:
        return False
    return None


def extract_tags(value) -> list[str]:
    """All well-formed technique IDs in a cell, uppercased, order-preserving."""
    seen, tags = set(), []
    for match in TAG_RE.findall(str(value or "")):
        tag = match.upper()
        if tag not in seen:
            seen.add(tag)
            tags.append(tag)
    return tags


def _detect_columns(rows: list[list]) -> tuple[int, dict]:
    """(header row index, {field: column index}) for the best-scoring row in
    the scan window; ({}, -1) when no row yields a name column."""
    best_idx, best_map = -1, {}
    for idx, row in enumerate(rows[:_HEADER_SCAN_ROWS]):
        mapping = {}
        for col, cell in enumerate(row):
            header = _norm(cell)
            if not header:
                continue
            for field, synonyms in COLUMN_SYNONYMS.items():
                if field not in mapping and header in synonyms:
                    mapping[field] = col
                    break
        if "name" in mapping and len(mapping) > len(best_map):
            best_idx, best_map = idx, mapping
    return best_idx, best_map


def _rows_to_use_cases(rows, header_idx, columns, ref_prefix, row_number):
    """Data rows below the header -> use-case row dicts. row_number maps a
    0-based grid index to the human row reference (1-based sheet row)."""

    def cell(row, field):
        col = columns.get(field)
        if col is None or col >= len(row):
            return None
        value = row[col]
        return None if value is None else str(value).strip() or None

    use_cases, skipped = [], 0
    for idx in range(header_idx + 1, len(rows)):
        row = rows[idx]
        if not any(str(v).strip() for v in row if v is not None):
            continue
        name = cell(row, "name")
        if not name:
            skipped += 1
            continue
        if len(use_cases) >= MAX_USE_CASE_ROWS:
            raise IngestError(
                f"Use-case file exceeds the {MAX_USE_CASE_ROWS:,}-row limit. "
                "Split the export and run separate assessments."
            )
        use_cases.append(
            {
                "row_ref": f"{ref_prefix}:{row_number(idx)}",
                "name": name[:500],
                "description": (cell(row, "description") or "")[:2000] or None,
                "log_source": (cell(row, "log_source") or "")[:255] or None,
                "enabled": parse_enabled(row[columns["enabled"]])
                if "enabled" in columns and columns["enabled"] < len(row)
                else None,
                "tags": extract_tags(cell(row, "tags")),
                "logic": cell(row, "logic"),
            }
        )
    return use_cases, skipped


def _xlsx_grids(content: bytes) -> list[tuple[str, list[list]]]:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise IngestError(f"Could not read the Excel file: {exc}") from exc
    grids = []
    total_rows = 0
    try:
        if len(wb.sheetnames) > MAX_SHEETS:
            raise IngestError(
                f"Workbook has more than {MAX_SHEETS} sheets — remove the "
                "extra sheets and retry."
            )
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                if len(row) > MAX_ROW_CELLS:
                    raise IngestError(
                        f"Sheet '{ws.title}' has rows wider than "
                        f"{MAX_ROW_CELLS} columns — this doesn't look like a "
                        "rule or asset export. Please use the template."
                    )
                total_rows += 1
                if total_rows > MAX_TOTAL_ROWS:
                    raise IngestError(
                        f"Workbook exceeds the {MAX_TOTAL_ROWS:,}-row overall "
                        "limit. Split the export and run separate assessments."
                    )
                rows.append(list(row))
            grids.append((ws.title, rows))
    finally:
        wb.close()
    return grids


def _xls_grids(content: bytes) -> list[tuple[str, list[list]]]:
    import xlrd

    try:
        wb = xlrd.open_workbook(file_contents=content)
    except Exception as exc:
        raise IngestError(f"Could not read the legacy .xls file: {exc}") from exc
    sheets = wb.sheets()
    if len(sheets) > MAX_SHEETS:
        raise IngestError(
            f"Workbook has more than {MAX_SHEETS} sheets — remove the extra "
            "sheets and retry."
        )
    grids, total_rows = [], 0
    for sheet in sheets:
        if sheet.ncols > MAX_ROW_CELLS:
            raise IngestError(
                f"Sheet '{sheet.name}' has rows wider than {MAX_ROW_CELLS} "
                "columns — this doesn't look like a rule or asset export. "
                "Please use the template."
            )
        total_rows += sheet.nrows
        if total_rows > MAX_TOTAL_ROWS:
            raise IngestError(
                f"Workbook exceeds the {MAX_TOTAL_ROWS:,}-row overall limit. "
                "Split the export and run separate assessments."
            )
        grids.append((sheet.name, [sheet.row_values(r) for r in range(sheet.nrows)]))
    return grids


def _csv_grid(content: bytes) -> list[list]:
    text = content.decode("utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    # Same workbook budgets the xlsx/xls readers enforce (2026-08-02
    # review nit — csv previously buffered unbounded rows/cells).
    rows = []
    for row in csv.reader(io.StringIO(text), dialect):
        if len(row) > MAX_ROW_CELLS:
            raise IngestError(
                f"The CSV has rows wider than {MAX_ROW_CELLS} columns — this "
                "doesn't look like a rule export. Please use the template."
            )
        rows.append(row)
        if len(rows) > MAX_TOTAL_ROWS:
            raise IngestError(
                f"The CSV exceeds the {MAX_TOTAL_ROWS:,}-row overall limit. "
                "Split the export and run separate assessments."
            )
    return rows


_SAMPLE_ROWS = 5      # data rows echoed in the parse preview (Phase 9 wizard)
_SAMPLE_CELL_CAP = 200


def _preview_cells(row, width: int) -> list:
    """Stringify + cap one grid row for the parse preview."""
    cells = ["" if v is None else str(v)[:_SAMPLE_CELL_CAP] for v in row[:width]]
    return cells + [""] * (width - len(cells))


def validate_column_override(override, header_row) -> dict:
    """Explicit {field: column_index} map from the wizard -> validated map.
    Raises IngestError (-> 422) on unknown fields, non-integer or
    out-of-range indexes, duplicate targets, or a missing name column."""
    if not isinstance(override, dict) or not override:
        raise IngestError("Column mapping must be a non-empty object of field: column index.")
    width = len(header_row)
    validated = {}
    for field, idx in override.items():
        if field not in COLUMN_SYNONYMS:
            raise IngestError(
                f"Unknown field '{field}'. Valid fields: {', '.join(sorted(COLUMN_SYNONYMS))}."
            )
        if isinstance(idx, bool) or not isinstance(idx, int) or not 0 <= idx < width:
            raise IngestError(
                f"Column index for '{field}' is out of range — this file has "
                f"{width} columns."
            )
        validated[field] = idx
    if "name" not in validated:
        raise IngestError("A use-case name column is required.")
    if len(set(validated.values())) != len(validated):
        raise IngestError("Each column can only be mapped to one field.")
    return validated


def parse_use_case_file(content: bytes, file_type: str, column_override: dict | None = None) -> dict:
    """Parse a use-case dump into rows + detected column map.

    column_override (Phase 9 wizard): an explicit {field: 0-based column
    index} map that REPLACES the auto-detected mapping — the header row and
    sheet are still located by detection, only the field→column assignment
    is overridden (validated against the header width).

    Returns {"rows": [{row_ref, name, description, log_source, enabled,
    tags[], logic}], "columns": {field: 0-based index}, "sheet": str|None,
    "headers": [str], "sample_rows": [[str]], "warnings": [str],
    "row_count": int}. Raises IngestError (-> 422).
    """
    if file_type in ("pdf", "docx"):
        raise IngestError(PDF_DOCX_MESSAGE)
    if not content or not content.strip():
        raise IngestError("The uploaded file is empty.")

    warnings = []
    if file_type == "csv":
        candidates = [("csv", _csv_grid(content))]
    elif file_type == "xlsx":
        candidates = _xlsx_grids(content)
    elif file_type == "xls":
        candidates = _xls_grids(content)
    else:
        raise IngestError(f"Unsupported use-case file type: {file_type}")

    for sheet_name, rows in candidates:
        header_idx, columns = _detect_columns(rows)
        if header_idx < 0:
            continue
        if column_override is not None:
            columns = validate_column_override(column_override, rows[header_idx])
        ignored = [name for name, _ in candidates if name != sheet_name]
        if ignored:
            warnings.append(
                "Only sheet '%s' was parsed; ignored: %s" % (sheet_name, ", ".join(ignored))
            )
        if "tags" not in columns:
            warnings.append(
                "No MITRE technique column detected — every rule will need AI "
                "tagging. If your file has one, use “Adjust columns” to map it."
            )
        use_cases, skipped = _rows_to_use_cases(
            rows, header_idx, columns, sheet_name, lambda i: i + 1
        )
        if skipped:
            warnings.append(f"{skipped} row(s) skipped: no value in the name column.")
        if not use_cases:
            raise IngestError(
                "No use-case rows found below the detected header row."
            )
        headers = _preview_cells(rows[header_idx], len(rows[header_idx]))
        sample_rows = [
            _preview_cells(row, len(headers))
            for row in rows[header_idx + 1 :]
            if any(str(v).strip() for v in row if v is not None)
        ][:_SAMPLE_ROWS]
        return {
            "rows": use_cases,
            "columns": columns,
            "sheet": sheet_name,
            "headers": headers,
            "sample_rows": sample_rows,
            "warnings": warnings,
            "row_count": len(use_cases),
        }

    raise IngestError(NO_NAME_COLUMN_MESSAGE)


# ---------------------------------------------------------------------------
# Environment workbook
# ---------------------------------------------------------------------------

SHEET_SYNONYMS = {
    "assets": {
        "assets", "asset", "platforms", "assets platforms", "asset platforms",
        "environment", "asset inventory",
        "asset list", "infrastructure", "environment inventory", "systems",
        "platforms in scope", "in scope assets", "inventory",
    },
    "log_sources": {
        "log sources", "log source", "logsources", "logs", "data sources",
        "telemetry", "onboarded log sources",
        "log inventory", "onboarded sources", "siem sources", "sources",
        "log source inventory", "data onboarding", "telemetry sources",
    },
    "tooling": {
        "tooling", "security tooling", "tools", "security tools", "security stack",
        "security products", "toolset", "defensive tooling",
        "security controls", "controls", "tool inventory",
    },
    "crown_jewels": {
        "crown jewels", "crown jewel", "crownjewels", "critical assets",
        "critical services",
        "key assets", "business critical assets", "high value assets", "hva",
        "critical systems", "crown jewel assets",
    },
}

# Ordered longest-synonym-first at module load so "cisco ios" wins over "ios"
# and "azure ad" over "azure". Canonical names = ATT&CK v19.1 platform
# vocabulary (enterprise + mobile; ICS techniques carry no real platforms).
_PLATFORM_RULES = sorted(
    [
        ("windows", "Windows"), ("win server", "Windows"), ("win", "Windows"),
        ("linux", "Linux"), ("ubuntu", "Linux"), ("rhel", "Linux"),
        ("red hat", "Linux"), ("centos", "Linux"), ("debian", "Linux"),
        ("suse", "Linux"), ("fedora", "Linux"),
        ("macos", "macOS"), ("mac os", "macOS"), ("os x", "macOS"),
        ("osx", "macOS"), ("mac", "macOS"),
        ("containers", "Containers"), ("container", "Containers"),
        ("kubernetes", "Containers"), ("k8s", "Containers"),
        ("docker", "Containers"), ("openshift", "Containers"),
        ("aks", "Containers"), ("gke", "Containers"), ("eks", "Containers"),
        ("ecs", "Containers"), ("rancher", "Containers"),
        ("containerd", "Containers"), ("podman", "Containers"),
        ("esxi", "ESXi"), ("vmware", "ESXi"), ("vsphere", "ESXi"),
        ("vcenter", "ESXi"),
        ("iaas", "IaaS"), ("aws", "IaaS"), ("amazon web services", "IaaS"),
        ("ec2", "IaaS"), ("gcp", "IaaS"), ("google cloud", "IaaS"),
        ("azure", "IaaS"), ("oci", "IaaS"), ("oracle cloud", "IaaS"),
        ("openstack", "IaaS"), ("digitalocean", "IaaS"),
        ("digital ocean", "IaaS"), ("alibaba cloud", "IaaS"),
        ("saas", "SaaS"), ("salesforce", "SaaS"), ("servicenow", "SaaS"),
        ("slack", "SaaS"), ("github", "SaaS"), ("workday", "SaaS"),
        ("zoom", "SaaS"), ("atlassian", "SaaS"), ("jira", "SaaS"),
        ("confluence", "SaaS"), ("gitlab", "SaaS"),
        ("office 365", "Office Suite"), ("o365", "Office Suite"),
        ("m365", "Office Suite"), ("microsoft 365", "Office Suite"),
        ("office suite", "Office Suite"), ("google workspace", "Office Suite"),
        ("gsuite", "Office Suite"), ("g suite", "Office Suite"),
        ("exchange online", "Office Suite"), ("sharepoint", "Office Suite"),
        ("microsoft teams", "Office Suite"), ("outlook", "Office Suite"),
        ("onedrive", "Office Suite"),
        ("identity provider", "Identity Provider"), ("idp", "Identity Provider"),
        ("azure ad", "Identity Provider"), ("aad", "Identity Provider"),
        ("entra id", "Identity Provider"), ("entra", "Identity Provider"),
        ("okta", "Identity Provider"), ("ping identity", "Identity Provider"),
        ("active directory", "Identity Provider"),
        ("adfs", "Identity Provider"), ("keycloak", "Identity Provider"),
        ("jumpcloud", "Identity Provider"), ("duo", "Identity Provider"),
        ("onelogin", "Identity Provider"), ("ping federate", "Identity Provider"),
        ("network devices", "Network Devices"), ("network device", "Network Devices"),
        ("network infrastructure", "Network Devices"), ("network", "Network Devices"),
        ("router", "Network Devices"), ("routers", "Network Devices"),
        ("switch", "Network Devices"), ("switches", "Network Devices"),
        ("firewall", "Network Devices"), ("firewalls", "Network Devices"),
        ("cisco ios", "Network Devices"), ("juniper", "Network Devices"),
        ("vpn", "Network Devices"), ("load balancer", "Network Devices"),
        ("f5", "Network Devices"), ("palo alto", "Network Devices"),
        ("fortigate", "Network Devices"), ("fortinet", "Network Devices"),
        ("check point", "Network Devices"), ("checkpoint", "Network Devices"),
        ("meraki", "Network Devices"), ("aruba", "Network Devices"),
        ("sonicwall", "Network Devices"), ("mikrotik", "Network Devices"),
        ("android", "Android"),
        ("ios", "iOS"), ("iphone", "iOS"), ("ipad", "iOS"),
        ("ipados", "iOS"),
    ],
    key=lambda rule: len(rule[0]),
    reverse=True,
)

_ICS_MARKERS = {
    "ot", "ics", "scada", "plc", "plcs", "dcs", "hmi", "industrial",
    "industrial control", "industrial control systems",
    "operational technology", "ot ics", "ot assets", "ics assets",
    "iiot", "modbus",
}
# NOTE: no "android"/"ios" here — those are platform rules (so "Cisco IOS"
# resolves to Network Devices, not a mobile fleet); the platform branch sets
# the mobile flag for real Android/iOS entries.
_MOBILE_MARKERS = {
    "mobile", "mobile fleet", "managed mobile", "mdm", "intune", "jamf",
    "workspace one", "mobileiron", "mobile devices",
    "airwatch", "kandji", "soti", "maas360", "emm",
}


def _word_match(needle: str, haystack: str) -> bool:
    return bool(re.search(rf"(?<![a-z0-9]){re.escape(needle)}(?![a-z0-9])", haystack))


def _match_platform(entry_norm: str) -> str | None:
    for synonym, platform in _PLATFORM_RULES:
        if _word_match(synonym, entry_norm):
            return platform
    return None


def _sheet_entries(rows: list[list], skipped: list | None = None) -> list[str]:
    """First-column entries of a sheet, honoring an optional second
    'present?' column (a recognizable False skips the row) and skipping a
    header-ish first row. Phase 14g: rows skipped via Present=No are also
    recorded into ``skipped`` when a list is passed (evidence trail)."""
    entries = []
    for idx, row in enumerate(rows):
        first = str(row[0]).strip() if row and row[0] is not None else ""
        if not first:
            continue
        if idx == 0 and _norm(first) in {
            "platform", "platforms", "asset", "assets", "name", "log source",
            "log sources", "source", "tool", "tooling", "crown jewel",
            "crown jewels", "entry", "item",
        }:
            continue
        if len(row) > 1 and parse_enabled(row[1]) is False:
            if skipped is not None:
                skipped.append(first)
            continue
        entries.append(first)
        if len(entries) > MAX_ASSET_ROWS:
            raise IngestError(
                f"Environment workbook exceeds the {MAX_ASSET_ROWS:,}-row limit."
            )
    return entries


def parse_environment_file(content: bytes, file_type: str) -> dict:
    """Parse the multi-sheet environment workbook.

    Returns {"environment": <Phase 0 applicability contract, exclusions
    empty — intake merges them>, "log_sources": [...], "tooling": [...],
    "crown_jewels": [...], "sheets_found": {...}, "assumptions": [str],
    "warnings": [str]}. Raises IngestError (-> 422).
    """
    if file_type not in ("xlsx", "xls"):
        raise IngestError(
            "The environment workbook must be an Excel file (.xlsx/.xls) — "
            "use the ScopeWise environment template (sheets: Assets, Log "
            "Sources, Security Tooling, Crown Jewels)."
        )
    if not content or not content.strip():
        raise IngestError("The uploaded environment workbook is empty.")

    grids = _xlsx_grids(content) if file_type == "xlsx" else _xls_grids(content)

    sheets_found: dict = {}
    for sheet_name, rows in grids:
        normalized = _norm(sheet_name)
        for kind, synonyms in SHEET_SYNONYMS.items():
            if kind not in sheets_found and normalized in synonyms:
                sheets_found[kind] = (sheet_name, rows)
                break

    assumptions, warnings = [], []
    platforms, unmatched = [], []
    has_ics, has_mobile = False, False
    # Phase 14g: per-entry evidence trail — additive output only, no
    # semantic change to platforms/flags/lists.
    interpretations: list[dict] = []

    if "assets" in sheets_found:
        skipped_assets: list = []
        for entry in _sheet_entries(sheets_found["assets"][1], skipped=skipped_assets):
            # Markers and platform matching are independent: "OT/SCADA
            # network" both flags ICS and maps to Network Devices.
            entry_norm = _norm(entry)
            matched = False
            read_as = []
            if any(_word_match(m, entry_norm) for m in _ICS_MARKERS):
                has_ics = True
                matched = True
                read_as.append("enabled the ICS/OT matrix")
            if any(_word_match(m, entry_norm) for m in _MOBILE_MARKERS):
                has_mobile = True
                matched = True
                read_as.append("enabled the Mobile matrix")
            platform = _match_platform(entry_norm)
            if platform:
                if platform in ("Android", "iOS"):
                    has_mobile = True
                if platform not in platforms:
                    platforms.append(platform)
                matched = True
                read_as.append(f"counted as platform {platform}")
            if not matched:
                unmatched.append(entry)
            interpretations.append({
                "entry": entry,
                "sheet": "Assets",
                "interpretation": "; ".join(read_as)
                or "not recognized — ignored for platform filtering",
            })
        interpretations.extend(
            {"entry": e, "sheet": "Assets",
             "interpretation": "skipped — you marked it Present = No"}
            for e in skipped_assets
        )
        if unmatched:
            shown = ", ".join(unmatched[:10])
            more = f" (+{len(unmatched) - 10} more)" if len(unmatched) > 10 else ""
            assumptions.append(
                f"asset entries not mapped to ATT&CK platforms (ignored for "
                f"platform filtering): {shown}{more}"
            )
    else:
        assumptions.append(
            "environment workbook has no recognizable Assets/Platforms sheet — "
            "platform filtering skipped; coverage % is a lower bound"
        )

    _LIST_INTERPRETATIONS = {
        "log_sources": "log source onboarded — feeds detection feasibility "
                       "(short-term roadmap bucketing)",
        "tooling": "security tooling you own — feeds mid-term feasibility "
                   "(telemetry you could onboard)",
        "crown_jewels": "noted as a crown-jewel asset",
    }

    def _plain_list(kind: str, label: str) -> list[str]:
        if kind in sheets_found:
            skipped: list = []
            entries = _sheet_entries(sheets_found[kind][1], skipped=skipped)
            interpretations.extend(
                {"entry": e, "sheet": label,
                 "interpretation": _LIST_INTERPRETATIONS[kind]}
                for e in entries
            )
            interpretations.extend(
                {"entry": e, "sheet": label,
                 "interpretation": "skipped — you marked it Present = No"}
                for e in skipped
            )
            return entries
        assumptions.append(
            f"environment workbook has no recognizable {label} sheet"
        )
        return []

    log_sources = _plain_list("log_sources", "Log Sources")
    tooling = _plain_list("tooling", "Security Tooling")
    crown_jewels = _plain_list("crown_jewels", "Crown Jewels")

    return {
        "interpretations": interpretations,
        "environment": {
            "platforms": platforms,
            "has_ics_assets": has_ics,
            "has_managed_mobile": has_mobile,
            "inventory_provided": "assets" in sheets_found,
            "exclusions": [],
        },
        "log_sources": log_sources,
        "tooling": tooling,
        "crown_jewels": crown_jewels,
        "sheets_found": {k: v[0] for k, v in sheets_found.items()},
        "assumptions": assumptions,
        "warnings": warnings,
    }
